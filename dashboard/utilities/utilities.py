import uuid  
import pytz
from ..constants import constants
from openpyxl.styles import Font, PatternFill, Border, Side
from datetime import date, datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment
import calendar
import os
from openpyxl.drawing.image import Image
from ..models import (Activity, User)
import numpy as np
from openpyxl.utils import get_column_letter
from collections import defaultdict
from io import BytesIO
from dateutil.relativedelta import relativedelta

def generate_unique_id():
    new_id = "EMP_" + str(uuid.uuid1())
    return new_id

# Helper function to convert and format a datetime to Cairo timezone
def convert_to_cairo_timezone_and_format(datetime_obj):
    cairo_timezone = pytz.timezone('Africa/Cairo')
    return datetime_obj.astimezone(cairo_timezone).strftime("%Y-%m-%d %H:%M:%S")

def format_hr_codes(userGrade):
    from dashboard.models import User
    users = User.objects.filter(grade=userGrade).order_by('hrCode')
    # Iterate over the users and update their HR codes
    new_hr_code = 1
    for user in users:
        user_hr_code = f'{constants.USER_GRADE_CHOICES[userGrade][1]}{new_hr_code:03d}'
        # Check if the HR code has changed
        if user.hrCode != user_hr_code:
            user.hrCode = user_hr_code
            user.save()

        new_hr_code += 1

def check_is_admin(id):
    from dashboard.models import User
    user = User.objects.filter(id=id).first()

    # if not super user dont show users
    if not (user.is_superuser or user.isAdmin):
        return False
    return True

def convert_grade_to_choice(grade):
    grade_mappings = {
        'A1': constants.GRADE_A_1,
        'A2': constants.GRADE_A_2,
        'A3': constants.GRADE_A_3,
        'B1': constants.GRADE_B_1,
        'B2': constants.GRADE_B_2,
        'B3': constants.GRADE_B_3,
        'B4': constants.GRADE_B_4,
        'B5': constants.GRADE_B_5
    }
    return grade_mappings.get(grade, None)

def convert_expert_to_choice(expert_string):
    expert_mappings = {
        'EXP': constants.EXPERT_USER,
        'LOC': constants.LOCAL_USER
    }
    return expert_mappings.get(expert_string, None)

def convert_nat_group_to_choice(nat_group_value):
    if isinstance(nat_group_value, str):
        nat_group_string = nat_group_value.strip().upper()
        nat_group_mappings = {
            'VCH': constants.VCH,
            'FUD': constants.FUD,
            'TD': constants.TD,
            'CWD': constants.CWD,
            'PSD': constants.PSD,
            'RSMD': constants.RSMD
        }
        return nat_group_mappings.get(nat_group_string, None)
    elif isinstance(nat_group_value, int):
        return nat_group_value
    else:
        return None  # Handle other cases if needed

def convert_company_to_choice(company_string):
    company_mappings = {
        'OCG': constants.OCG,
        'NK': constants.NK,
        'EHAF': constants.EHAF,
        'ACE': constants.ACE,
        'TD': constants.TD
    }
    return company_mappings.get(company_string, None)

def generate_username_from_name(name, taken_usernames):
    # Split the name and create a base username
    name_parts = name.split()
    base_username = f"{name_parts[0]}_{name_parts[-1]}".lower()

    # Initialize a counter
    counter = 0
    unique_username = base_username

    # Check if the username exists in the taken usernames set and find a unique one by appending a number
    while unique_username in taken_usernames:
        counter += 1
        unique_username = f"{base_username}_{counter}"

    # Add the new unique username to the taken usernames set
    taken_usernames.add(unique_username)

    return unique_username

def generate_first_name(name):
    if name:
        return name.split()[0]
    return None

def generate_last_name(name):
    if name:
        return ' '.join(name.split()[1:])
    return None

def __add_local_working_days__(current_date, user, cover_ws):
    start_date = current_date.date().replace(day=1)
    last_day_of_month = calendar.monthrange(current_date.date().year, current_date.date().month)[1]
    end_date = (current_date.date() + timedelta(days=last_day_of_month)).replace(day=1)

    # Adjust end_date to the next day
    end_date = end_date + timedelta(days=1)

    total_working_days_expert = constants.JAPAN_WORKING_DAYS[current_date.month]
    cover_ws['H35'].value = total_working_days_expert
    cover_ws['H35'].alignment = Alignment(horizontal='center', vertical='center')  # Center the text

    start_date = current_date.date().replace(day=1)
    last_day_of_month = calendar.monthrange(current_date.date().year, current_date.date().month)[1]
    end_date = current_date.date().replace(day=last_day_of_month)

    last_day_of_month = calendar.monthrange(current_date.date().year, current_date.date().month)[1]
    end_date = current_date.date().replace(day=last_day_of_month) + timedelta(days=1)
    total_working_days = np.busday_count(start_date, end_date, weekmask='1111111')
    # Iterate through each day in the range
    for day in range(1, last_day_of_month + 1):
        day_off = Activity.objects.filter(
            user=user,
            activityDate__day=current_date.date().day,
            activityDate__month=current_date.date().month,
            activityDate__year=current_date.date().year,
            activityType=constants.OFFDAY,
        ).exists()

        if day_off and (day == 5 and day - 1 in (4, 6) and day + 1 in (4, 6)):
            total_working_days -= 1

        # Filter activities for the current user and month excluding 'H' type activities
        activities = Activity.objects.filter(
            user=user,
            activityDate__month=current_date.date().month,
            activityDate__year=current_date.date().year,
        ).filter(activityType__in=[constants.INCAIRO, constants.HOLIDAY])

        # Count the number of activities
        working_days = activities.count()
        # Iterate over weeks in the month
        for week in calendar.monthcalendar(current_date.year, current_date.month):
            for day in week:
                # Check if Thursday and Saturday are off-days in the current week
                if day != 0:  # Ignore days that belong to the previous or next month (represented as 0)
                    thursday_offday = any(activity.activityType == constants.OFFDAY for activity in activities.filter(activityDate__day=day + 3))
                    saturday_offday = any(activity.activityType == constants.OFFDAY for activity in activities.filter(activityDate__day=day + 5))

                    # If either Thursday or Saturday is off-day, decrement working_days
                    if thursday_offday or saturday_offday:
                        working_days -= 1

        # Add "0" under "Cairo" for local users
        cover_ws['F35'].value = working_days
        cover_ws['F35'].alignment = Alignment(horizontal='center', vertical='center')  # Center the text
        cover_ws['J35'].value = total_working_days
        cover_ws['J35'].alignment = Alignment(horizontal='center', vertical='center')  # Center the text

        start_date = current_date.date().replace(day=1)
        end_date = current_date.date().replace(day=calendar.monthrange(current_date.date().year, current_date.date().month)[1])
        end_date = end_date + relativedelta(days=1)

        activities_japan = Activity.objects.filter(
            user=user,
            activityDate__month=current_date.date().month,
            activityDate__year=current_date.date().year,
        ).filter(activityType__in=[constants.HOMEASSIGN]).count()

        cover_ws['D35'].value = activities_japan
        cover_ws['D35'].alignment = Alignment(horizontal='center', vertical='center')  # Center the text

        formula = "=ROUND(F35/J35, 3)"
        cover_ws["N35"].value = formula
        cover_ws['N35'].font = Font(size=11)
        cover_ws['N35'].alignment = Alignment(horizontal='center', vertical='center')

         # Japan NOD/TCD
        formula = "=ROUND(C35/F35, 3)"
        cover_ws['L35'].value = formula
        cover_ws['L35'].font = Font(size=11)
        cover_ws['L35'].alignment = Alignment(horizontal='center', vertical='center')

def __add_expert_working_days__(current_date, user, cover_ws):
        ###### LOCAL ######
        start_date = current_date.date().replace(day=1)
        last_day_of_month = calendar.monthrange(current_date.date().year, current_date.date().month)[1]
        end_date = current_date.date().replace(day=last_day_of_month)

        last_day_of_month = calendar.monthrange(current_date.date().year, current_date.date().month)[1]
        end_date = current_date.date().replace(day=last_day_of_month) + timedelta(days=1)
        total_working_days_cairo = np.busday_count(start_date, end_date, weekmask='1111111')
         ###### LOCAL ######
        start_date = current_date.date().replace(day=1)
        last_day_of_month = calendar.monthrange(current_date.date().year, current_date.date().month)[1]
        end_date = (current_date.date() + timedelta(days=last_day_of_month)).replace(day=1)

        # Adjust end_date to the next day
        end_date = end_date + timedelta(days=1)
        total_working_days_japan = constants.JAPAN_WORKING_DAYS[current_date.month]

        # Filter activities for the current user and month excluding 'H' type activities
        activities = Activity.objects.filter(
            user=user,
            activityDate__month=current_date.date().month,
            activityDate__year=current_date.date().year,
        ).filter(activityType__in=[constants.HOMEASSIGN])

        # Count the number of activities
        working_days = activities.count()

        cover_ws['D35'].value = working_days
        cover_ws['D35'].alignment = Alignment(horizontal='center', vertical='center')  # Center the text

        cover_ws['H35'].value = total_working_days_japan
        cover_ws['H35'].alignment = Alignment(horizontal='center', vertical='center')  # Center the text

        start_date = current_date.date().replace(day=1)
        end_date = current_date.date().replace(day=calendar.monthrange(current_date.date().year, current_date.date().month)[1])
        end_date = end_date + relativedelta(days=1)

        activities_cairo_count = Activity.objects.filter(
            user=user,
            activityDate__month=current_date.date().month,
            activityDate__year=current_date.date().year,
            activityType__in=[constants.INCAIRO, constants.HOLIDAY]
        ).count()

        # Cairo NOD
        cover_ws['F35'].value = activities_cairo_count
        cover_ws['F35'].alignment = Alignment(horizontal='center', vertical='center')  # Center the text

        # Cairo TCD
        cover_ws['J35'].value = total_working_days_cairo
        cover_ws['J35'].alignment = Alignment(horizontal='center', vertical='center')  # Center the text

        # Japan NOD/TCD
        formula = "=ROUND(D35/H35, 3)"
        cover_ws['L35'].value = formula
        cover_ws['L35'].font = Font(size=11)
        cover_ws['L35'].alignment = Alignment(horizontal='center', vertical='center')

        # Cairo NOD/TCD
        formula = "=ROUND(F35/J35, 3)"
        cover_ws['N35'].value = formula
        cover_ws['N35'].font = Font(size=11)
        cover_ws['N35'].alignment = Alignment(horizontal='center', vertical='center')

def set_borders(ws, row, columns):
    for col in columns:
        cell_address = f'{col}{row}'
        ws[cell_address].border = Border(top=Side(style='thin', color='000000'),
                                        left=Side(style='thin', color='000000'),
                                        right=Side(style='thin', color='000000'),
                                        bottom=Side(style='thin', color='000000'))

def set_other_borders(ws, row, columns):
    for col in columns:
        cell_address = f'{col}{row}'
        ws[cell_address].border = Border(right=Side(style='thin', color='000000'),
                                        bottom=Side(style='thin', color='000000'),
                                        left=Side(style='thin', color='000000'),
                                        top=Side(style='thin', color='000000'))

def __format_cell__(cell, value, size=11, italic=False, center=True):
    cell.value = value
    cell.font = Font(size=size, italic=True, bold=True)
    if center:
        cell.alignment = Alignment(horizontal='center', vertical='center')

def __add_daily_activities_sheet__(wb, current_date, user, counter):
    # Extract month and year
    month = current_date.month
    year = current_date.year

    # get all the activities for that month by that user
    activities = Activity.objects.filter(activityDate__month=month, activityDate__year=year, user__id=user.id).order_by('-created')
    user = User.objects.get(id=user.id)

    daily_activities = wb.create_sheet(title=f"{str(counter)} (DA)")

    dateFont = Font(size=16)

    name_year_month_border = Border(
    left=Side(border_style='thin'),
    right=Side(border_style='thin'),
    top=Side(border_style='thin'),
    bottom=Side(border_style='thin')
)

    # Merge cells for the new section
    daily_activities.merge_cells(start_row=2, start_column=2, end_row=5, end_column=10)
    for row in range(2, 6):
        for col in range(2, 11):
            cell = daily_activities.cell(row=row, column=col)
            cell.border = name_year_month_border

    # Set the text and alignment in the first cell of the merged range
    merged_cell = daily_activities.cell(row=2, column=2)
    merged_cell.value = "Consulting Services for Greater Cairo Metro Line No. 4 Phase1 Project"
    merged_cell.font = dateFont
    merged_cell.alignment = Alignment(horizontal='center', vertical='center')

    script_directory = os.path.dirname(os.path.abspath(__file__))
    parent_directory = os.path.dirname(script_directory)
    parent_parent_directory = os.path.dirname(parent_directory)
    logo_path = os.path.join(parent_parent_directory, 'static', 'images', 'logo.png')
    img = Image(logo_path)
    img.height = 1.08 * 72  # 1 inch = 72 points
    img.width = 1.14 * 72
    daily_activities.add_image(img, 'M2')

    # Set font and border for "Year:" label in cell A8
    year_label_cell = daily_activities.cell(row=8, column=1, value="Year:")
    year_label_cell.font = Font(bold=True, size=12)
    year_label_cell.border = name_year_month_border

    # Align "Year:" label horizontally and vertically
    year_label_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Set font and border for the year value in cell B8
    year_value_cell = daily_activities.cell(row=8, column=2, value=year)
    year_value_cell.font = Font(bold=True, size=12)
    year_value_cell.border = name_year_month_border

    # Align year value horizontally and vertically
    year_value_cell.alignment = Alignment(horizontal='center', vertical='center')
    year_digits = year % 100
    month_name = calendar.month_name[month]

    # Combine the month name and the last two digits of the year
    formatted_month = f"{month_name}-{year_digits:02d}"

    cell = daily_activities.cell(row=9, column=2, value=formatted_month)
    cell.border = name_year_month_border
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='center', vertical='center')  # Center the text horizontally and vertically

    # Set the text "Month" in cell A9
    cell_month = daily_activities.cell(row=9, column=1, value="Month")
    cell_month.border = name_year_month_border
    cell_month.font = Font(bold=True, size=12)
    cell_month.alignment = Alignment(horizontal='center', vertical='center')  # Center the text horizontally and vertically

    # Adjust the width of column A to fit the text
    daily_activities.column_dimensions['A'].width = 10
    daily_activities.column_dimensions['B'].width = 10 # Adjust the height as needed

    daily_activities.merge_cells(start_row=8, start_column=7, end_row=8, end_column=11)
    daily_activities.merge_cells(start_row=9, start_column=7, end_row=9, end_column=11)
    daily_activities.merge_cells(start_row=12, start_column=5, end_row=12, end_column=14)

    for row in range(12, 13):
        for col in range(5, 15):
            cell = daily_activities.cell(row=row, column=col)
            cell.border = name_year_month_border

    cell = daily_activities.cell(row=8, column=7, value=user.get_full_name())
    cell.font = Font(bold=True, size=12)

    # Set alignment
    cell.alignment = Alignment(horizontal='center', vertical='center')

    # Create border
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    # Apply border to each cell in the merged range
    for row in daily_activities.iter_rows(min_row=8, max_row=8, min_col=7, max_col=11):
        for cell in row:
            cell.border = border

    cell = daily_activities.cell(row=9, column=7, value=user.position)
    cell.font = Font(bold=True, size=12)

    # Set alignment
    cell.alignment = Alignment(horizontal='center', vertical='center')

    # Create border
    border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))

    # Apply border to each cell in the merged range
    for row in daily_activities.iter_rows(min_row=9, max_row=9, min_col=7, max_col=11):
        for cell in row:
            cell.border = border

    # Create headers for columns
    headers_border = Border(left=Side(style='medium'), 
            right=Side(style='medium'), 
            top=Side(style='medium'), 
            bottom=Side(style='medium'))
    
    for col in range(1, 4):
        header_cell = daily_activities.cell(row=12, column=col, value=["Day", "Cairo", "Japan"][col - 1])
        header_cell.font = dateFont
        header_cell.border = headers_border
        header_cell.alignment = Alignment(vertical='center')

    cell = daily_activities.cell(row=12, column=5, value="DAILY ACTIVITIES")
    cell.font = dateFont
    cell.border = name_year_month_border

    # Center the text horizontally and vertically
    cell.alignment = Alignment(horizontal='center', vertical='center')

    for day in range(1, calendar.monthrange(year, month)[1] + 1):
        start_row_index = day + 12  # Offset by 12 to account for existing rows]

        cell = daily_activities.cell(row=start_row_index, column=1)
        cell.value = f"{day:02d}"
        cell.font = Font(size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Set borders for each cell within the merged range
        for col in ['A', 'B', 'C']:
            for row in range(start_row_index, start_row_index + 1):
                cell_address = f'{col}{row}'
                daily_activities[cell_address].border = Border(top=Side(style='thin', color='000000'),
                                                                left=Side(style='thin', color='000000'),
                                                                right=Side(style='thin', color='000000'),
                                                                bottom=Side(style='thin', color='000000'))

        # Fetch activities for the current day
        activities_for_day = activities.filter(activityDate__day=day)

        # Combine activities' text and type
        activities_text = "\n".join([activity.userActivity or '' for activity in activities_for_day])
        activities_type = "\n".join([str(activity.get_activity_type()) for activity in activities_for_day])

        # Calculate the number of merged rows for activities
        num_merged_rows = activities_text.count('\n') + 1

        for col in range(5, 14):
            bottom_cell_address = f"{get_column_letter(col)}{start_row_index}"
            daily_activities[bottom_cell_address].border = Border(bottom=Side(style='thin', color='000000'))

        merged_range = f"E{start_row_index}:N{start_row_index}"  # Update the range accordingly
        daily_activities.merge_cells(merged_range)
        merged_cell = daily_activities.cell(row=start_row_index, column=5)  # Top-left cell of the merged range
        merged_cell.value = activities_text
        merged_cell.font = Font(size=10)
        merged_cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')

        font_size = 10
        line_height = 1.2 * font_size
        row_height_factor = 2
        required_height = line_height * num_merged_rows * row_height_factor

        # Adjust the row height for each merged row
        for i in range(start_row_index, start_row_index + num_merged_rows):
            row_dimension = daily_activities.row_dimensions[i]
            row_dimension.height = required_height

        # Set the activities type in the appropriate column
        if user.expert in [constants.LOCAL_USER, constants.EXPERT_USER]:
            cell = daily_activities.cell(row=start_row_index, column=3 if activities_type == "J" else 2)
            cell.value = activities_type
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(size=11)

def __add_cover_sheet__(wb, current_month_name, current_year, user, current_date, current_month, counter):
    cover_ws = wb.create_sheet(title=str(counter))
    cover_ws.page_setup.print_scaling = 90
    # Set print area to end at S54
    cover_ws.print_area = 'A1:A2'
    
    # Adjust page margins
    cover_ws.page_margins.left = 0.5
    cover_ws.page_margins.right = 0.5
    cover_ws.page_margins.top = 0.5
    cover_ws.page_margins.bottom = 0.5

    # Adjust scaling to fit the content to fewer pages
    cover_ws.page_setup.fitToWidth = 1
    cover_ws.page_setup.fitToHeight = 0

    # Merge and set the title
    border_style = Border(
        bottom=Side(border_style='thin'),
    )

    # add the logo
    script_directory = os.path.dirname(os.path.abspath(__file__))
    parent_directory = os.path.dirname(script_directory)
    parent_parent_directory = os.path.dirname(parent_directory)
    logo_path = os.path.join(parent_parent_directory, 'static', 'images', 'logo.png')
    img = Image(logo_path)
    img.height = 1.08 * 72  # 1 inch = 72 points
    img.width = 1.14 * 72
    # Add the image to the worksheet
    cover_ws.add_image(img, 'O2')

    title_cell = cover_ws['A1']
    cover_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=13)
    title_cell.value = constants.COVER_TS_TEXT
    title_cell.font = Font(size=11)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Merge and set the Monthly Time Sheet text
    cover_ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=13)  # Modified here
    monthly_ts_cell = cover_ws['A3']  # Modified here
    monthly_ts_cell.value = 'Monthly Time Sheet'
    monthly_ts_cell.font = Font(size=16, italic=True, bold=True)
    monthly_ts_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Set the month and year
    year_label_cell = cover_ws.cell(row=5, column=1, value="Month:")
    year_label_cell.font = Font(bold=True, size=12)

    year_label_cell = cover_ws.cell(row=5, column=11, value="Year:")
    year_label_cell.font = Font(bold=True, size=12)
    year_label_cell.alignment = Alignment(horizontal='left')

    cover_ws.merge_cells(start_row=5, start_column=13, end_row=5, end_column=14)  # Modified here
    year_label_cell = cover_ws.cell(row=5, column=13, value=current_year)
    year_label_cell.font = Font(bold=True, size=10)

    cover_ws.merge_cells(start_row=5, start_column=3, end_row=5, end_column=6)  # Modified here
    value_year_cell = cover_ws.cell(row=5, column=3, value=current_month_name)
    value_year_cell.font = Font(bold=True, size=12)
    
    for column in range(3, 7):  # Columns 3 to 6 inclusive
        cell = cover_ws.cell(row=5, column=column)
        cell.border = border_style

    name_label_cell = cover_ws.cell(row=7, column=1, value="Name:")
    name_label_cell.font = Font(bold=True, size=12)

    cover_ws.merge_cells(start_row=7, start_column=3, end_row=7, end_column=10)  # Modified here
    name_label_cell = cover_ws.cell(row=7, column=3, value=str(user.get_full_name()))
    name_label_cell.font = Font(bold=True, size=12)

    for column in range(3, 11):  # Columns 3 to 10 inclusive
        cell = cover_ws.cell(row=7, column=column)
        cell.border = border_style

    for column in range(13, 15):  # Columns 3 to 10 inclusive
        cell = cover_ws.cell(row=5, column=column)
        cell.border = border_style

    user_grade = user.get_grade()
    name_label_cell = cover_ws.cell(row=9, column=1, value="Category:")
    name_label_cell.font = Font(bold=True, size=12)

    cover_ws.cell(row=9, column=4, value="A1")
    cover_ws.cell(row=9, column=7, value="A2")
    cover_ws.cell(row=9, column=10, value="A3")

    cover_ws.cell(row=11, column=4, value="B1")
    cover_ws.cell(row=11, column=7, value="B2")
    cover_ws.cell(row=11, column=10, value="B3")

    cover_ws.cell(row=13, column=4, value="B4")
    cover_ws.cell(row=13, column=7, value="B5")

    # Define the rows and columns
    rows = [9, 11, 13]
    columns = [4, 7, 10]

    # Iterate over each cell and set the value and alignment
    for row in rows:
        for col in columns:
            cell = cover_ws.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center')


    # Mapping of grades to corresponding columns
    grade_mapping = {
        'A1': 'C9',
        'A2': 'F9',
        'A3': 'I9',
        'B1': 'C11',
        'B2': 'F11',
        'B3': 'I11',
        'B4': 'C13',
        'B5': 'F13',
    }

    # # Set the letter '/' in the corresponding cell based on the user's grade
    if user_grade in grade_mapping:
        grade_column = grade_mapping[user_grade]
        cell_address = grade_column
        cover_ws[cell_address].value = '✓'
        cover_ws[cell_address].alignment = Alignment(horizontal='center', vertical='center')

    # Set borders for various cells
    set_other_borders(cover_ws, 9, ['C', 'F', 'I'])
    set_other_borders(cover_ws, 11, ['C', 'F', 'I'])
    set_other_borders(cover_ws, 13, ['C', 'F'])

    cell = cover_ws.cell(row=15, column=1, value="Nationality:")
    cell.font = Font(bold=True, size=12)

    cell = cover_ws.cell(row=15, column=6, value="Expatriate")
    cell.font = Font(name='Times New Roman', size=11)
    cell.alignment = Alignment(horizontal='center', vertical='center')

    cell = cover_ws.cell(row=15, column=10, value="Local")
    cell.font = Font(name='Times New Roman', size=11)
    cell.alignment = Alignment(horizontal='center', vertical='center')

    # Get the user's nationality
    user_nationality = user.get_expert()

    # Mapping of nationalities to corresponding columns
    nationality_mapping = {
        'EXP': 'D',  # Column for Expatriate
        'LOC': 'H',  # Column for Local
    }

    # Set the letter '/' in the corresponding cell and center it
    if user_nationality in nationality_mapping:
        nationality_column = nationality_mapping[user_nationality]
        cell_address = f'{nationality_column}15'
        cover_ws[cell_address].value = '✓'
        cover_ws[cell_address].alignment = Alignment(horizontal='center', vertical='center')

    set_other_borders(cover_ws, 15, ['D', 'H'])

    cell = cover_ws.cell(row=17, column=1, value="Group Field:")
    cell.font = Font(bold=True, size=12)

    cover_ws.merge_cells(start_row=17, start_column=5, end_row=17, end_column=9)
    cover_ws.merge_cells(start_row=17, start_column=11, end_row=17, end_column=17)
    cover_ws.merge_cells(start_row=19, start_column=5, end_row=19, end_column=9)
    cover_ws.merge_cells(start_row=19, start_column=11, end_row=19, end_column=17)

    cell = cover_ws.cell(row=17, column=5, value="Management and SHQE")
    cell.font = Font(size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center')

    cell = cover_ws.cell(row=17, column=11, value="Tender Evaluation and Contract Negotiation")
    cell.font = Font(size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center')

    cell = cover_ws.cell(row=19, column=5, value="Construction Supervision")
    cell.font = Font(size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center')

    cell = cover_ws.cell(row=19, column=11, value="O&M")
    cell.font = Font(size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center')

    set_other_borders(cover_ws, 17, ['D', 'J'])
    set_other_borders(cover_ws, 19, ['D', 'J'])

    grey_fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
    first_day_of_month = datetime(current_year, current_month, 1)
    shift_amount = 1
    for i in range(1, 17):
        col_address = chr(ord('A') + (i - 1 + shift_amount))  # Alternating columns C and G
        # Set the day of the week
        day_of_week_cell = cover_ws[col_address + '24']
        day_of_week_cell.value = (first_day_of_month + timedelta(days=i - 1)).strftime("%a")
        day_of_week_cell.alignment = Alignment(horizontal='center', vertical='center')
        day_of_week_cell.fill = grey_fill
        set_borders(cover_ws, 24, [col_address])
        set_borders(cover_ws, 26, [col_address])
        set_borders(cover_ws, 27, [col_address])
        set_borders(cover_ws, 28, [col_address])
        set_borders(cover_ws, 29, [col_address])
        # Set the day of the month
        day_of_month_cell = cover_ws[col_address + '25']
        day_of_month_cell.value = str((first_day_of_month + timedelta(days=i - 1)).day)
        day_of_month_cell.alignment = Alignment(horizontal='center', vertical='center')
        day_of_month_cell.fill = grey_fill
        set_borders(cover_ws, 25, [col_address])
        cell_address_activity_type = f'{col_address}26'

        # Replace 'activity_date' with the actual date for which you want to retrieve the activityType
        activity_date = first_day_of_month + timedelta(days=i - 1)
        activity_instance = Activity.objects.filter(user=user, activityDate=activity_date).first()

        if activity_instance:
            activity_type = activity_instance.get_activity_type()
            cover_ws[cell_address_activity_type].value = str(activity_type)
            if activity_type == constants.ACTIVITY_TYPES_CHOICES[constants.HOLIDAY][1]:
                cover_ws[cell_address_activity_type].fill = grey_fill
        else:
            cover_ws[cell_address_activity_type].value = None
        cover_ws[cell_address_activity_type].alignment = Alignment(horizontal='center', vertical='center')

    first_day_of_second_half = datetime(current_year, current_month, 17).date()

    # Write abbreviated weekdays in row 30 for the second half of the month
    for i in range(0, calendar.monthrange(current_year, current_month)[1] - 16):
        col_address = chr(ord('A') + i + shift_amount) + '27'
        day_of_week = (first_day_of_second_half + timedelta(days=i)).strftime("%a")  # Get the abbreviated day name
        cell = cover_ws[col_address]
        cell.value = day_of_week
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = grey_fill
        set_borders(cover_ws, 27, [col_address])
        cell.border = Border(top=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'), left=Side(style='thin', color='000000'))

    # Continue writing the numeric values for the remaining days in the row below (row 30)
    for i in range(17, calendar.monthrange(current_year, current_month)[1] + 1):
        set_borders(cover_ws, 26, [col_address])
        col_address = chr(ord('A') + (i - 17 + shift_amount))
        cell_address = f'{col_address}28'
        cell = cover_ws[cell_address]
        cell.value = (first_day_of_month + timedelta(days=i - 1)).day
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = grey_fill
        set_borders(cover_ws, 28, [col_address])
        set_borders(cover_ws, 29, [col_address])

    for i in range(17, calendar.monthrange(current_year, current_month)[1] + 1):
        col_address_activity_type = chr(ord('A') + (i - 17 + shift_amount)) + '29'

        activity_date = datetime(current_year, current_month, 17) + timedelta(days=i - 17)
        activity_instance = Activity.objects.filter(user=user, activityDate=activity_date).first()

        cell = cover_ws[col_address_activity_type]
        if activity_instance:
            activity_type = activity_instance.get_activity_type()
            cell.value = str(activity_type)
            if activity_type == constants.ACTIVITY_TYPES_CHOICES[constants.HOLIDAY][1]:
                cell.fill = grey_fill
        else:
            cell.value = None
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row in range(33, 36):
        for col_letter in ['D', 'F', 'H', 'J', 'L', 'N']:
            cell = cover_ws[col_letter + str(row)]
            cell.border = Border(
                left=Side(style='thin', color='000000' if col_letter != 'C' else '000000'),
                right=Side(style='thin', color='000000' if col_letter != 'J' else '000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )

    # add summary for working days
    cover_ws.merge_cells('D32:G33')
    cover_ws['D32'].value = "No. of Days (NOD)*"
    cover_ws['D32'].font = Font(size=11, bold=True)
    cover_ws['D32'].alignment = Alignment(horizontal='center', vertical='center')  # Center the text
    cover_ws['D32'].fill = grey_fill

    # Assuming user is an instance of the User model
    if user.expert == constants.LOCAL_USER:
        __add_local_working_days__(current_date, user, cover_ws)
    elif user.expert == constants.EXPERT_USER:
        __add_expert_working_days__(current_date, user, cover_ws)

    # Merge cells and set values for total calendar days (TCD)
    cover_ws.merge_cells('H32:K33')
    tcd_cell = cover_ws['H32']
    tcd_cell.value = "Total Calendar Days\n(TCD)"
    tcd_cell.font = Font(size=10, bold=True)
    tcd_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    tcd_cell.fill = grey_fill

    # Set values for Japan and Cairo columns
    cover_ws.merge_cells('D34:E34')
    cover_ws.merge_cells('D35:E35')
    cover_ws.merge_cells('F34:G34')
    cover_ws.merge_cells('F35:G35')
    cover_ws.merge_cells('H34:I34')
    cover_ws.merge_cells('H35:I35')
    cover_ws.merge_cells('J34:K34')
    cover_ws.merge_cells('J35:K35')

    labels = {
        'D34': 'Japan',
        'F34': 'Cairo',
        'H34': 'Japan',
        'J34': 'Cairo'
    }

    for cell_address, label_text in labels.items():
        cell = cover_ws[cell_address]
        cell.value = label_text
        cell.font = Font(size=11, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Merge cells and set values for consumption NOD/TCD
    cover_ws.merge_cells('L32:O33')
    consumption_cell = cover_ws['L32']
    consumption_cell.value = "Consumption NOD/TCD"
    consumption_cell.font = Font(size=11, bold=True)
    consumption_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    consumption_cell.fill = grey_fill

    # # Set values for Japan and Cairo columns in consumption NOD/TCD section
    cover_ws.merge_cells('L34:M34')
    cover_ws.merge_cells('N34:O34')
    cover_ws.merge_cells('L35:M35')
    cover_ws.merge_cells('N35:O35')

    labels = {
        'L34': 'Japan',
        'N34': 'Cairo'
    }
    for cell_address, label_text in labels.items():
        cell = cover_ws[cell_address]
        cell.value = label_text
        cell.font = Font(size=11, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    cover_ws['G33'].border = Border(right=Side(style='thin', color='000000'))
    cover_ws['G32'].border = Border(right=Side(style='thin', color='000000'))
    cover_ws['D33'].border = Border(left=Side(style='thin', color='000000'))
    cover_ws['D32'].border = Border(left=Side(style='thin', color='000000'))

    cover_ws['K33'].border = Border(right=Side(style='thin', color='000000'))
    cover_ws['K32'].border = Border(right=Side(style='thin', color='000000'))
    cover_ws['H33'].border = Border(left=Side(style='thin', color='000000'))
    cover_ws['H32'].border = Border(left=Side(style='thin', color='000000'))

    cover_ws['O33'].border = Border(right=Side(style='thin', color='000000'))
    cover_ws['O32'].border = Border(right=Side(style='thin', color='000000'))
    cover_ws['L33'].border = Border(left=Side(style='thin', color='000000'))
    cover_ws['L32'].border = Border(left=Side(style='thin', color='000000'))

    for char in ['D', 'E', 'F', 'G', "H", "I", "J", "K", "L", "M", "N", "O"]:
        cover_ws[f'{char}32'].border = Border(top=Side(style='thin', color='000000'), left=Side(style='thin', color='000000'), right=Side(style="thin", color="000000"))

    __format_cell__(cover_ws['C38'], "NOCE Approval")

    # NAT Approval cell (L38)
    __format_cell__(cover_ws['L38'], "NAT Approval")

    cover_ws.merge_cells(start_row=38, start_column=3, end_row=38, end_column=6)
    cover_ws.merge_cells(start_row=38, start_column=12, end_row=38, end_column=15)

    for col_letter in range(ord('A'), ord('S')):
        col_letter = chr(col_letter)
        cell = cover_ws[col_letter + '42']
        cell.border = Border(bottom=Side(style='thick'))

    for column in range(3, 7):  # Columns 3 to 10 inclusive
        cell = cover_ws.cell(row=41, column=column)
        cell.border = border_style

    for column in range(12, 16):  # Columns 3 to 10 inclusive
        cell = cover_ws.cell(row=41, column=column)
        cell.border = border_style

    cover_ws.merge_cells("B43:F43")
    cover_ws.merge_cells("B44:F44")
    cover_ws.merge_cells("M43:Q43")
    labels = {
        'B43': "J = Working day In Japan",
        'M43': "C = Working day In Cairo",
        'B44': "H = Official Holiday In Cairo",
        'M44': "X = Day off",
        'B46': "Note: According to the contract 81/M the total days are working days in Cairo plus to official holiday in Egypt\n *NOD=C (Working day in Cairo)+H (Official Holiday in Egypt)"
    }

    cover_ws.merge_cells(start_row=46, start_column=2, end_row=48, end_column=17)
    cell = cover_ws['B46']
    cell.value = labels['B46']

    # Enable text wrapping
    cell.alignment = Alignment(wrap_text=True)

    for cell_address, label_text in labels.items():
        cover_ws.merge_cells(f'{cell_address}:{chr(ord(cell_address[0]) + 2)}{cell_address[1:]}')
        cell = cover_ws[cell_address]
        cell.value = label_text
        cell.font = Font(size=11, bold=True)

    # Set column width to 4.5
    for col_letter in range(ord('A'), ord('S')):
        col_letter = chr(col_letter)
        cover_ws.column_dimensions[col_letter].width = 4.5

    for row in range(48, cover_ws.max_row + 1):
        cover_ws.row_dimensions[row].hidden = True

    # Hide columns beyond S
    for col in range(19, cover_ws.max_column + 1):
        cover_ws.column_dimensions[get_column_letter(col)].hidden = True
    # Set the height of all rows to 15 points
    for row in cover_ws.iter_rows():
        for cell in row:
            cover_ws.row_dimensions[cell.row].height = 15
        
    for row in range(48, 50):  # Unhide rows 48 and 49
        cover_ws.row_dimensions[row].hidden = False


def create_activity_excel_report(users, activities, selected_date, companyName, date):
    date = datetime.combine(date, datetime.min.time())

    current_date = date
    current_month = current_date.month
    current_year = current_date.year

    font = Font(size=8)

    # Create a Border object for cell borders
    thin_border = Border(top=Side(style='thin'),
                        bottom=Side(style='thin'),
                        left=Side(style='medium'),  # Make the left border thick
                        right=Side(style='medium'),)
    # Create a fill (background color) for the headers
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Get the current month and year
    current_month_name = current_date.strftime("%B")

    last_day_of_month = (current_date.replace(day=1, month=current_month % 12 + 1, year=current_year) - timedelta(days=1)).day
    day_headers = [str(day) for day in range(1, last_day_of_month + 1)]

    constants.EXPORT_ACTIVITY_COLUMNS += ["Japan", "Cairo", "Japan %", "Cairo %"]
    constants.EXPORT_ACTIVITY_COLUMNS += day_headers
    addition_headers = ["Dep NAT", "Invoiced"]
    constants.EXPORT_ACTIVITY_COLUMNS += addition_headers

    # Create a new Excel workbook and add a worksheet
    wb = Workbook()

    counter = 1
    # Create individual timesheet/daily activities for every user
    for user in users:
        __add_cover_sheet__(wb, current_month_name, current_year, user, current_date, current_month, counter)
        __add_daily_activities_sheet__(wb, current_date, user, counter)
        counter += 1

    ws = wb.active
    ws.title = "TS"

    ws.cell(row=1, column=1, value="Cairo Line 4 Phase 1 - TimeSheet")
    # Set the font to bold
    bold_font = Font(bold=True)
    ws.cell(row=1, column=1).font = bold_font
    # Combine the month abbreviation and year in the desired format
    formatted_date = f"{current_month_name.lower()[:3]}-{current_year}"
    ws.cell(row=1, column=5, value=formatted_date)
    yearFont = Font(size=8)
    ws.cell(row=1, column=5).font = yearFont
    ws.cell(row=1, column=5).fill = yellow_fill

    # Create a mapping dictionary for user types
    for col_num, column_title in enumerate(constants.EXPORT_ACTIVITY_COLUMNS, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = column_title
        cell.alignment = Alignment(horizontal="center")
        cell.font = font if column_title.isnumeric() else None
        # Apply thick border above the headers
        cell.border = Border(outline=Side(style='medium'))

    for col_num in range(1, len(constants.EXPORT_ACTIVITY_COLUMNS) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.border = Border(bottom=Side(style='medium'))
        cell = ws.cell(row=3, column=col_num)
        cell.border = Border(bottom=Side(style='medium'))

    # Initialize a counter variable for rows
    current_row = 2
    # Use a defaultdict for activities
    default_activities = {day: '' for day in range(1, last_day_of_month + 1)}

    user_rows = {}
    for user in users:
        user_id = str(user.id)
        user_row = user_rows[user_id] = {
        'user_counter': current_row,
        'full_name': user.get_full_name(),
        'company': user.get_company(),
        'position': user.position,
        'expert': user.get_expert(),
        'grade': user.get_grade(),
        'nat_group': user.get_natGroup(),
        'invoiced': 'X',
        'Japan': '',
        'Cairo': '',
        'Japan %': '',
        'Cairo %': '',
        'activities': defaultdict(str, default_activities.copy())
    }
        # Retrieve activities for the current user
        user_activities = activities.filter(user=user)
        for activity in user_activities:
            day = activity.activityDate.day
            activity_type = activity.get_activity_type()
            user_row['activities'][day] = activity_type

        current_row += 1

    for user_id, user_data in user_rows.items():
        row_num = user_data['user_counter'] + 2
        ws.cell(row=row_num, column=1, value=row_num - 3).font = font  # Add userCounter
        ws.cell(row=row_num, column=2, value=user_data['full_name']).font = Font(size=8, bold=True)
        ws.cell(row=row_num, column=3, value=str(user_data['company'])).font = font
        ws.cell(row=row_num, column=4, value=str(user_data['position'])).font = font
        ws.cell(row=row_num, column=5, value=str(user_data['expert'])).font = font
        ws.cell(row=row_num, column=6, value=str(user_data['grade'])).font = font
        ws.cell(row=row_num, column=7 + last_day_of_month + 5, value=str(user_data['nat_group'])).font = font
        ws.cell(row=row_num, column=7 + last_day_of_month + 6, value=str(user_data['invoiced'])).font = font

        cairo_count = 0
        total_working_days_cairo = 0
        japan_count = 0
        total_working_days_japan = 0
        if user_data['expert'] == constants.EXPERT_LOCAL_CHOICES[constants.EXPERT_USER][1]:
            start_date = current_date.replace(day=1)
            end_date = current_date.replace(day=calendar.monthrange(current_date.year, current_date.month)[1])
            end_date = end_date + relativedelta(days=1)
            # Convert start_date and end_date to datetime64[D]
            start_date = np.datetime64(start_date, 'D')
            end_date = np.datetime64(end_date, 'D')

            total_working_days_cairo = np.busday_count(start_date, end_date, weekmask='1111111')
            total_working_days_japan = constants.JAPAN_WORKING_DAYS[current_date.month]
            # np.busday_count(start_date, end_date, weekmask='0011111')

        elif user_data['expert'] == constants.EXPERT_LOCAL_CHOICES[constants.LOCAL_USER][1]:
            start_date = current_date.replace(day=1)
            last_day_of_month = calendar.monthrange(current_date.year, current_date.month)[1]
            end_date = current_date.replace(day=last_day_of_month)
            last_day_of_month = calendar.monthrange(current_date.year, current_date.month)[1]
            end_date = current_date.replace(day=last_day_of_month) + timedelta(days=1)
            start_date = np.datetime64(start_date, 'D')
            end_date = np.datetime64(end_date, 'D')
            total_working_days_cairo = np.busday_count(start_date, end_date, weekmask='1111111')
            total_working_days_japan = constants.JAPAN_WORKING_DAYS[current_date.month]

        for col, activity_type in user_data['activities'].items():
            if (activity_type == 'J'):
                japan_count += 1
            if activity_type in ['C', 'H']:
                cairo_count += 1

            ws.cell(row=row_num, column=col + 11, value=str(activity_type)).font = font
            if "C" in activity_type:
                green_fill = PatternFill(start_color="A8D08D", end_color="A8D08D", fill_type="solid")
                ws.cell(row=row_num, column=col + 11).fill = green_fill
            if "X" in activity_type:
                red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                ws.cell(row=row_num, column=col + 11).fill = red_fill
            if "H" in activity_type:
                grey_fill = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")
                ws.cell(row=row_num, column=col + 11).fill = grey_fill
            if "J" in activity_type:
                pink_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
                ws.cell(row=row_num, column=col + 11).fill = pink_fill

        # Check if total_working_days_cairo is not zero before calculating the percentage
        cairo_percentage = round(cairo_count / total_working_days_cairo, 5) if total_working_days_cairo != 0 else 0

        # Check if total_working_days_japan is not zero before calculating the percentage
        japan_percentage = round(japan_count / total_working_days_japan, 5) if total_working_days_japan != 0 else 0

        # worked days
        red_bold_italic_font = Font(size=12, color="FF0000", bold=True, italic=True)
        ws.cell(row=row_num, column=7 + 2, value=cairo_count).font = red_bold_italic_font
        ws.cell(row=row_num, column=7 + 1, value=japan_count).font = red_bold_italic_font

        ws.cell(row=row_num, column=7 + 4, value=cairo_percentage).font = red_bold_italic_font
        ws.cell(row=row_num, column=7 + 3, value=japan_percentage).font = red_bold_italic_font

        for cell in ws[row_num]:
            cell.border = thin_border
        ws.freeze_panes = ws.cell(row=4, column=8)

    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        ws.column_dimensions[col[0].column_letter].width = max_length

    # Center align all cells in column #
    for cell in ws['A']:
        cell.alignment = Alignment(horizontal='center')

    # Create AutoFilter for all columns starting from row 2
    ws.auto_filter.ref = f'A2:{get_column_letter(ws.max_column)}{ws.max_row}'

    for col_idx in range(12, 12 + last_day_of_month):
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = 3

    # Create sheets for each Dep NAT
    dep_nats = list(set(user.get_natGroup() for user in users if user.get_natGroup() is not None))
    for dep_nat in dep_nats:
        if dep_nat is not None:
            sheet = wb.create_sheet(title=str(dep_nat))  # Create a new sheet for each Dep NAT
            sheet.sheet_properties.tabColor = "FFA500"
            __customize_sheet__(sheet, dep_nat, selected_date)
            # Write "Dep Nat" values under the appropriate column
    for user_id, user_data in user_rows.items():
        if user_data['nat_group'] == dep_nat:
            row_num = user_data['user_counter'] + 2
            dep_nat_cell = sheet.cell(row=row_num, column=constants.EXPORT_ACTIVITY_COLUMNS.index("Dep NAT") + 1)
            dep_nat_cell.value = str(dep_nat)
            dep_nat_cell.font = Font(size=8)
            dep_nat_cell.alignment = Alignment(horizontal="center")

    excel_data = BytesIO()
    wb.save(excel_data)
    excel_data.seek(0)

    excel_data = BytesIO()
    wb.save(excel_data)
    excel_data.seek(0)
    from django.http import HttpResponse
    response = HttpResponse(excel_data, content_type="application/ms-excel")
    response["Content-Disposition"] = f'attachment; filename=timesheet.xlsx'
    return response

def __department_mapping__(dep_nat):
    group = constants.DEP_MAPPING.get(dep_nat, 'Unknown')
    return group

def __customize_sheet__(sheet, dep_nat, selected_date):
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.sheet_view.showGridLines = False

    # Add a logo at the top right
    script_directory = os.path.dirname(os.path.abspath(__file__))
    parent_directory = os.path.dirname(script_directory)
    parent_parent_directory = os.path.dirname(parent_directory)
    logo_path = os.path.join(parent_parent_directory, 'static', 'images', 'logo.png')
    img = Image(logo_path)
    sheet.add_image(img, 'H1')

    # Create a reverse mapping from dep_nat to the corresponding number
    nat_group_mapping_reverse = {name: number for number, name in constants.NAT_GROUP_CHOICES}
    # Get the number for the current dep_nat
    dep_nat_number = nat_group_mapping_reverse.get(dep_nat, -1)

    department = __department_mapping__(dep_nat_number)

    users = User.objects.filter(natGroup=dep_nat_number)

    # Set the title in the middle
    title_cell = sheet.cell(row=5, column=1, value=f"Dep NAT: {department}")
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal="center")

    # Merge cells for the title
    sheet.merge_cells(start_row=5, start_column=1, end_row=5, end_column=8)
    selected_date = datetime.strptime(selected_date, '%Y-%m-%d')
    month_name = calendar.month_name[selected_date.month]
    # Set the selected date below the title
    date_cell = sheet.cell(row=6, column=2, value=f"{month_name}")
    date_cell.font = Font(size=14)
    date_cell.alignment = Alignment(horizontal="center")

    # Merge cells for 'No.' and 'Name'
    sheet.merge_cells(start_row=8, start_column=1, end_row=8, end_column=4)
    sheet.merge_cells(start_row=8, start_column=5, end_row=8, end_column=8)

    # Set headers 'No.' and 'Name'
    sheet.cell(row=8, column=1, value='No.').font = Font(size=12, bold=True, color='000000')
    sheet.cell(row=8, column=5, value='Name').font = Font(size=12, bold=True, color='000000')

    # Add counter under 'No.' and extract names under 'Name'
    counter = 1
    for row_num, user in enumerate(users, start=9):
        sheet.cell(row=row_num, column=1, value=counter)
        sheet.cell(row=row_num, column=5, value=user.get_full_name())
        counter += 1

    # Hide cell lines surrounding the data of activities
    for row in sheet.iter_rows(min_row=9, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = None

    sheet.freeze_panes = sheet.cell(row=9, column=1)

