from .models import User, Activity, hrHistory
from rest_framework.response import Response
from rest_framework import status
from .serializers import CreateUserSerializer,\
      ListUsersSerializer, UserDeleteSerializer,\
      ActivitySerializer, CreateActivitySerializer,\
    MakeUserAdminSerializer, ChangePasswordSerializer,\
    UserTimeSheetSerializer, EditUserSerializer,\
        CalculateActivitySerializer, EditActivitySerializer
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from .constants import constants
from rest_framework_simplejwt.serializers import TokenObtainPairSerializer
from rest_framework_simplejwt.views import TokenObtainPairView
from django.shortcuts import get_object_or_404
from django.contrib.auth.hashers import make_password
from django.contrib.auth.tokens import default_token_generator
from datetime import date, datetime, timedelta
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill
from io import BytesIO
from django.http import HttpResponse
import pandas as pd
from .utilities import utilities
from openpyxl.styles import Font, Border, Side
from django.db import transaction
from openpyxl.drawing.image import Image
import os
import calendar
import numpy as np
from django.utils import timezone
import calendar
from dashboard.jobs.jobs import generate_noce_timesheet
from dateutil.relativedelta import relativedelta
from rest_framework.pagination import PageNumberPagination


class MyTokenObtainPairSerializer(TokenObtainPairSerializer):
    @classmethod
    def get_token(cls, user):
        token = super().get_token(user)
        token['username'] = user.username
        token['isAdmin'] = user.isAdmin
        token['is_superuser'] = user.is_superuser
        token['calendarType'] = user.calendarType
        return token
class MyTokenObtainPairView(TokenObtainPairView):
    serializer_class = MyTokenObtainPairSerializer

class UserRegistrationViewSet(viewsets.ViewSet):
    def create_user(self, request):
        data = request.data
        data['password'] = make_password(data['password'])  # Hash the password

        serializer = CreateUserSerializer(data=data)

        if serializer.is_valid():
            user = serializer.save()

            # Create a token for the user
            token = default_token_generator.make_token(user)

            return Response({
                # 'user': serializer.data,
                'token': token  # Include the token in the response
            }, status=status.HTTP_201_CREATED)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()

    def get_serializer_class(self):
        if self.action == 'create' or self.action == 'update':
            return CreateUserSerializer  # Replace with the appropriate serializer
        else:
            return ListUsersSerializer  # Replace with the serializer for other actions

    @action(detail=False, methods=['GET'])
    def get_users(self, request, *args, **kwargs):
        userId = request.user.id
        user = User.objects.filter(id=userId).first()

        # if not super user dont show users
        if not (user.is_superuser or user.isAdmin):
            return Response(status=status.HTTP_403_FORBIDDEN)
        users = User.objects.filter(isDeleted=False).order_by('created')
        serializer = self.get_serializer(users, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['GET'])
    def get_user(self, request, *args, **kwargs):
        userId = request.user.id
        user = User.objects.filter(id=userId).first()
        serializer = self.get_serializer(user)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['PATCH'])
    def make_admin(self, request, *args, **kwargs):
        adminId = request.user.id
        user = User.objects.filter(id=adminId).first()
        # if not super user dont show users
        if not user.is_superuser:
            return Response(status=status.HTTP_403_FORBIDDEN)

        serializer = MakeUserAdminSerializer(data=request.data)
        serializer.is_valid()
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        user_id = serializer.validated_data['userId']
        is_admin = serializer.validated_data['isAdmin']

        modified_user = User.objects.filter(id=user_id).first()
        modified_user.isAdmin = is_admin
        modified_user.save()
        return Response(status=status.HTTP_200_OK)

    @action(detail=False, methods=['PATCH'])
    def revoke_admin(self, request, *args, **kwargs):
        adminId = request.user.id
        user = User.objects.filter(id=adminId).first()
        # if not super user dont show users
        if not user.is_superuser:
            return Response(status=status.HTTP_403_FORBIDDEN)

        serializer = MakeUserAdminSerializer(data=request.data)
        serializer.is_valid()
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        user_id = serializer.validated_data['userId']

        modified_user = User.objects.filter(id=user_id).first()
        modified_user.isAdmin = False
        modified_user.save()
        return Response(status=status.HTTP_200_OK)

    # An endpoint for superusers only to edit any other user (grade, organizationcode..etc)
    @action(detail=False, methods=['patch'], url_path=r'edit_details/(?P<userId>\w+(?:-\w+)*)')
    def edit_details(self, request, *args, **kwargs):
        user_id = kwargs['userId']
        request_user = request.user

        if not request_user.is_superuser:
            return Response(status=status.HTTP_403_FORBIDDEN)

        user = User.objects.filter(id=user_id).first()
        if not user:
            return Response(status=status.HTTP_404_NOT_FOUND)

        serializer = EditUserSerializer(user, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(status=status.HTTP_200_OK)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['delete'])
    def delete_user(self, request, *args, **kwargs):
        adminId = request.user.id
        user = User.objects.filter(id=adminId).first()
        if not user.is_superuser:
            return Response(status=status.HTTP_403_FORBIDDEN)

        serializer = UserDeleteSerializer(data=request.data)  # Pass request data to the serializer

        if serializer.is_valid():  # Validate the data
            user_id = serializer.validated_data['userId']
            userObject = User.objects.filter(id=user_id).first()

            if not userObject:
                return Response(status=status.HTTP_400_BAD_REQUEST)
             # Check if the user to be deleted is a superuser
            if userObject.is_superuser:
                superusers_count = User.objects.filter(is_superuser=True).count()
                if superusers_count <= 1:
                    return Response(constants.LAST_SUPERUSER_DELETION_ERROR, status=status.HTTP_400_BAD_REQUEST)

            # Instead of deleting, mark the user as deleted
            userObject.isDeleted = True
            userObject.save()
            # Clear specific fields for the deleted user
            with transaction.atomic():
                userObject.grade = None
                userObject.hrCode = None
                userObject.organizationCode = None
                userObject.save()
            return Response(constants.SUCCESSFULLY_DELETED_USER, status=status.HTTP_200_OK)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['POST'])
    def reset_user_password(self, request, *args, **kwargs):
        adminId = request.user.id
        user = User.objects.filter(id=adminId).first()

         # if not super user dont show users
        if not user.is_superuser:
            return Response(status=status.HTTP_403_FORBIDDEN)
        serializer = UserDeleteSerializer(data=request.data)
        serializer.is_valid()
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        user_id = serializer.validated_data['userId']
        userObject = User.objects.filter(id=user_id).first()
        if not userObject:
            return Response(constants.CANT_RESET_USER_PASSWORD_ERROR, status=status.HTTP_400_BAD_REQUEST)
        userObject.password = make_password("1234")
        userObject.save()
        return Response(constants.SUCCESSFULLY_DELETED_USER, status=status.HTTP_200_OK)

    @transaction.atomic
    @action(detail=False, methods=['post'])
    def excel_sign_up(self, request, *args, **kwargs):
        """
        This endpoint creates bulk users by taking the first_name column and last_name column
        concatenate them to generate a username, and make a default password of 1234 that's required to be changed
        by the user
        """
        # Check if user is not an admin
        if not utilities.check_is_admin(request.user.id):
            return Response(status=status.HTTP_403_FORBIDDEN)
        # Check if an Excel file was uploaded
        if 'file' not in request.FILES:
            return Response({'detail': 'No file uploaded'}, status=status.HTTP_400_BAD_REQUEST)
        excel_file = request.FILES['file']

        unavailable_codes = hrHistory.objects.all().values_list('hrCode', flat=True)

        try:
            all_usernames = set(User.objects.all().values_list('username', flat=True))
            
            # Read the Excel file again, this time setting the second row as headers
            dataframe_with_headers = pd.read_excel(excel_file, header=2)
            # Find the column index for "HR Code" and "Remarks"
            hr_code_index = dataframe_with_headers.columns.get_loc("HR Code")
            remarks_index = dataframe_with_headers.columns.get_loc("Remarks") if "Remarks" in dataframe_with_headers.columns else None

            # If "Remarks" column is not found, we will display all columns up to the last one.
            if remarks_index is None:
                remarks_index = len(dataframe_with_headers.columns) - 1
            
            # Display the new column headers and a sample of the data to confirm
            dataframe_hr_to_remarks = dataframe_with_headers.iloc[:, hr_code_index:remarks_index+1]
            for index, row in dataframe_hr_to_remarks.iterrows():
                if pd.isna(row['Name']):
                    continue
                # Extract information from the row
                hr_code = row['HR Code']
                name = row['Name']
                email = row['Email Address']
                grade = row['Grade']
                organization_code = row['Organization Code']
                position = row['Position']
                department = row['Department']
                nat_group = row['NAT Group']
                working_location = row['Working Location']
                expert = row['Expert']
                mobilization_status = row['Mobilization status']
                company = row['Company']

                if hr_code in unavailable_codes:
                    return Response(constants.ERR_HR_CODE_USED, status=status.HTTP_400_BAD_REQUEST)

                # Assuming you have a function to convert the grade and expert to the corresponding integer choices
                grade_choice = utilities.convert_grade_to_choice(grade)
                expert_choice = utilities.convert_expert_to_choice(expert)
                nat_group_choice = utilities.convert_nat_group_to_choice(nat_group)
                company_choice = utilities.convert_company_to_choice(company)

                if expert_choice == constants.EXPERT_USER:
                    calendar_type = constants.CALENDAR_TYPE_EXPERT
                elif expert_choice == constants.LOCAL_USER:
                    calendar_type = constants.CALENDAR_TYPE_LOCAL
                else:
                    calendar_type = None

                email = row['Email Address'] if pd.notnull(row['Email Address']) else None
                new_username = utilities.generate_username_from_name(name, all_usernames)
                if User.objects.filter(username=new_username).exists() or User.objects.filter(email=email).exists():
                    # Skip this row if the username already exists
                    continue

                # Create a new User object
                user = User(
                    username = new_username,
                    email=email,
                    first_name=utilities.generate_first_name(name),
                    last_name=utilities.generate_last_name(name),
                    grade=grade_choice,
                    hrCode=hr_code,  # This will be overridden if grade is set and hrCode is not.
                    organizationCode=organization_code,
                    position=position,
                    department=department,
                    natGroup=nat_group_choice,
                    workingLocation=working_location,
                    expert=expert_choice,
                    company=company_choice,
                    mobilization=mobilization_status,
                    calendarType=calendar_type,
                    password=make_password("1234")
                )

                # Save the new user, the save method will call generate_hr_code if needed
                user.save()
            return Response({'message': 'Data extracted and printed successfully'}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def modify_details(self, request, *args, **kwargs):
        excel_file = request.FILES['file']
        modified_users = {} 
        try:
            dataframe_with_headers = pd.read_excel(excel_file, header=2)

            # Iterate over each row in the DataFrame
            for index, row in dataframe_with_headers.iterrows():
                # Extract username from the current row
                username = row['USER Name']
                # Check if a user with this username already exists
                existing_user = User.objects.filter(username=username).first()
                email = row['Email Address']
                organization_code = row['Organization Code']
                position = row['Position']
                department = row['Department']
                nat_group = row['NAT Group']
                nat_group_choice = utilities.convert_nat_group_to_choice(nat_group)
                working_location = row['Working Location']
                expert = row['Expert']
                expert_choice = utilities.convert_expert_to_choice(expert)
                mobilization_status = row['Mobilization status']
                company = row['Company']
                company_choice = utilities.convert_company_to_choice(company)

                name = row['Name']

                if existing_user:
                    modified_fields = []  # List to store modified fields for this user
                    if existing_user.email != email:
                        existing_user.email = email
                        modified_fields.append('email')
                    if existing_user.first_name != utilities.generate_first_name(name):
                        existing_user.first_name = utilities.generate_first_name(name)
                        modified_fields.append('first_name')
                    if existing_user.last_name != utilities.generate_last_name(name):
                        existing_user.last_name = utilities.generate_last_name(name)
                        modified_fields.append('last_name')
                    if existing_user.organizationCode != organization_code:
                        existing_user.organizationCode = organization_code
                        modified_fields.append('organizationCode')
                    if existing_user.position != position:
                        existing_user.position = position
                        modified_fields.append('position')
                    if existing_user.department != department:
                        existing_user.department = department
                        modified_fields.append('department')
                    if existing_user.natGroup != nat_group_choice:
                        existing_user.natGroup = nat_group_choice
                        modified_fields.append('natGroup')
                    if existing_user.workingLocation != working_location:
                        existing_user.workingLocation = working_location
                        modified_fields.append('workingLocation')
                    if existing_user.expert != expert_choice:
                        existing_user.expert = expert_choice
                        modified_fields.append('expert')
                    if existing_user.mobilization != mobilization_status:
                        existing_user.mobilization = mobilization_status
                        modified_fields.append('mobilization')
                    if existing_user.company != company_choice:
                        existing_user.company = company_choice
                        modified_fields.append('company')
                    
                    if modified_fields:  # Check if any fields were modified
                        modified_users[existing_user.username] = modified_fields  # Add to dictionary
                
                    existing_user.save()
                break
        except Exception as e:
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        # Return the dictionary of modified users and their modified fields
        return Response({'modified_users': modified_users}, status=status.HTTP_200_OK)

    @action(detail=False, methods=['PATCH'], url_path=r'change_password/(?P<userId>\w+(?:-\w+)*)')
    def change_password(self, request, *args, **kwargs):
        request_user_id = request.user.id
        target_user_id = kwargs['userId']
        # Check if the user accessing the endpoint is the same one that created the activity or not
        if str(request_user_id) != str(target_user_id):
            return Response(data=constants.NOT_ALLOWED_TO_ACCESS, status=status.HTTP_400_BAD_REQUEST)

        serializer = ChangePasswordSerializer(data=request.data)
        if serializer.is_valid():
            User.objects.filter(id=target_user_id).update(needsPasswordReset=False,\
                                                          password=make_password(serializer.validated_data['password']))
            return Response(data=constants.SUCCESSFULLY_CHANGED_PASSWORD, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    def __add_local_working_days_individual__(self, current_date, user, cover_ws):
        start_date = current_date.replace(day=1)
        last_day_of_month = calendar.monthrange(current_date.year, current_date.month)[1]
        end_date = (current_date + timedelta(days=last_day_of_month)).replace(day=1)

        # Adjust end_date to the next day
        end_date = end_date + timedelta(days=1)

        total_working_days_expert = constants.JAPAN_WORKING_DAYS[current_date.month]
        cover_ws['H35'].value = total_working_days_expert
        cover_ws['H35'].alignment = Alignment(horizontal='center', vertical='center')  # Center the text

        start_date = current_date.replace(day=1)
        last_day_of_month = calendar.monthrange(current_date.year, current_date.month)[1]
        end_date = current_date.replace(day=last_day_of_month)

        last_day_of_month = calendar.monthrange(current_date.year, current_date.month)[1]
        end_date = current_date.replace(day=last_day_of_month) + timedelta(days=1)
        total_working_days = np.busday_count(start_date, end_date, weekmask='1111111')

        # Iterate through each day in the range
        for day in range(1, last_day_of_month + 1):
            day_off = Activity.objects.filter(
                user=user,
                activityDate__day=current_date.day,
                activityDate__month=current_date.month,
                activityDate__year=current_date.year,
                activityType=constants.OFFDAY,
            ).exists()

            if day_off and (day == 5 and day - 1 in (4, 6) and day + 1 in (4, 6)):
                total_working_days -= 1

            # Filter activities for the current user and month excluding 'H' type activities
            activities = Activity.objects.filter(
                user=user,
                activityDate__month=current_date.month,
                activityDate__year=current_date.year,
            ).filter(activityType__in=[constants.INCAIRO, constants.HOLIDAY])

            # Count the number of activities
            working_days = activities.count()
            # Iterate over weeks in the month
            for week in calendar.monthcalendar(current_date.year, current_date.month):
                for day in week:
                    # Check if Thursday and Saturday are off-days in the current week
                    if day != 0:  # Ignore days that belong to the previous or next month (represented as 0)
                        thursday_offday = any(activity.activityType == constants.OFFDAY for activity in activities.filter(activityDate__day=day + 3))
                        saturday_offday = any(activity.activityType == constants.OFFDAY for activity in activities.filter(activityDate__day=day + 5))

                        # If either Thursday or Saturday is off-day, decrement working_days
                        if thursday_offday or saturday_offday:
                            working_days -= 1

            # Add "0" under "Cairo" for local users
            cover_ws['F35'].value = working_days
            cover_ws['F35'].alignment = Alignment(horizontal='center', vertical='center')  # Center the text
            cover_ws['J35'].value = total_working_days
            cover_ws['J35'].alignment = Alignment(horizontal='center', vertical='center')  # Center the text

            start_date = current_date.replace(day=1)
            end_date = current_date.replace(day=calendar.monthrange(current_date.year, current_date.month)[1])
            end_date = end_date + relativedelta(days=1)

            activities_japan = Activity.objects.filter(
                user=user,
                activityDate__month=current_date.month,
                activityDate__year=current_date.year,
            ).filter(activityType__in=[constants.HOMEASSIGN]).count()

            cover_ws['D35'].value = activities_japan
            cover_ws['D35'].alignment = Alignment(horizontal='center', vertical='center')  # Center the text

            formula = "=ROUND(F35/J35, 3)"
            cover_ws["N35"].value = formula
            cover_ws['N35'].font = Font(size=11)
            cover_ws['N35'].alignment = Alignment(horizontal='center', vertical='center')

             # Japan NOD/TCD
            formula = "=ROUND(C35/F35, 3)"
            cover_ws['L35'].value = formula
            cover_ws['L35'].font = Font(size=11)
            cover_ws['L35'].alignment = Alignment(horizontal='center', vertical='center')

    def __add_expert_working_days_individual__(self, current_date, user, cover_ws):
            ###### LOCAL ######
            start_date = current_date.replace(day=1)
            last_day_of_month = calendar.monthrange(current_date.year, current_date.month)[1]
            end_date = current_date.replace(day=last_day_of_month)

            last_day_of_month = calendar.monthrange(current_date.year, current_date.month)[1]
            end_date = current_date.replace(day=last_day_of_month) + timedelta(days=1)
            total_working_days_cairo = np.busday_count(start_date, end_date, weekmask='1111111')
            ###### LOCAL ######
            start_date = current_date.replace(day=1)
            last_day_of_month = calendar.monthrange(current_date.year, current_date.month)[1]
            end_date = (current_date + timedelta(days=last_day_of_month)).replace(day=1)

            # Adjust end_date to the next day
            end_date = end_date + timedelta(days=1)
            total_working_days_japan = constants.JAPAN_WORKING_DAYS[current_date.month]

            # Filter activities for the current user and month excluding 'H' type activities
            activities = Activity.objects.filter(
                user=user,
                activityDate__month=current_date.month,
                activityDate__year=current_date.year,
            ).filter(activityType__in=[constants.HOMEASSIGN])

            # Count the number of activities
            working_days = activities.count()

            cover_ws['D35'].value = working_days
            cover_ws['D35'].alignment = Alignment(horizontal='center', vertical='center')  # Center the text

            cover_ws['H35'].value = total_working_days_japan
            cover_ws['H35'].alignment = Alignment(horizontal='center', vertical='center')  # Center the text

            start_date = current_date.replace(day=1)
            end_date = current_date.replace(day=calendar.monthrange(current_date.year, current_date.month)[1])
            end_date = end_date + relativedelta(days=1)

            activities_cairo_count = Activity.objects.filter(
                user=user,
                activityDate__month=current_date.month,
                activityDate__year=current_date.year,
                activityType__in=[constants.INCAIRO, constants.HOLIDAY]
            ).count()

            # Cairo NOD
            cover_ws['F35'].value = activities_cairo_count
            cover_ws['F35'].alignment = Alignment(horizontal='center', vertical='center')  # Center the text

            # Cairo TCD
            cover_ws['J35'].value = total_working_days_cairo
            cover_ws['J35'].alignment = Alignment(horizontal='center', vertical='center')  # Center the text

            # Japan NOD/TCD
            formula = "=ROUND(D35/H35, 3)"
            cover_ws['L35'].value = formula
            cover_ws['L35'].font = Font(size=11)
            cover_ws['L35'].alignment = Alignment(horizontal='center', vertical='center')

            # Cairo NOD/TCD
            formula = "=ROUND(F35/J35, 3)"
            cover_ws['N35'].value = formula
            cover_ws['N35'].font = Font(size=11)
            cover_ws['N35'].alignment = Alignment(horizontal='center', vertical='center')

    def __add_individual_timesheet__(self, wb, current_month_name, current_year, user, current_date, current_month):
        cover_ws = wb.create_sheet(title=str(user.first_name))
        cover_ws.page_setup.print_scaling = 90
        cover_ws.print_area = 'A1:A2'

        # Adjust page margins
        cover_ws.page_margins.left = 0.5
        cover_ws.page_margins.right = 0.5
        cover_ws.page_margins.top = 0.5
        cover_ws.page_margins.bottom = 0.5

        # Adjust scaling to fit the content to fewer pages
        cover_ws.page_setup.fitToWidth = 1
        cover_ws.page_setup.fitToHeight = 0

        # Merge and set the title
        border_style = Border(
        bottom=Side(border_style='thin'),
        )

        # add the logo
        script_directory = os.path.dirname(os.path.abspath(__file__))
        # parent_directory = os.path.dirname(script_directory)
        parent_parent_directory = os.path.dirname(script_directory)
        logo_path = os.path.join(parent_parent_directory, 'static', 'images', 'logo.png')
        img = Image(logo_path)
        img.height = 1.08 * 72  # 1 inch = 72 points
        img.width = 1.14 * 72
        cover_ws.add_image(img, 'O2')
        cover_ws.column_dimensions['O'].width = 100  # Adjust the width as needed

        title_cell = cover_ws['A1']
        cover_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=13)
        title_cell.value = constants.COVER_TS_TEXT
        title_cell.font = Font(size=11)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Merge and set the Monthly Time Sheet text
        cover_ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=13)  # Modified here
        monthly_ts_cell = cover_ws['A3']  # Modified here
        monthly_ts_cell.value = 'Monthly Time Sheet'
        monthly_ts_cell.font = Font(size=16, italic=True, bold=True)
        monthly_ts_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Set the month and year
        year_label_cell = cover_ws.cell(row=5, column=1, value="Month:")
        year_label_cell.font = Font(bold=True, size=12)

        year_label_cell = cover_ws.cell(row=5, column=11, value="Year:")
        year_label_cell.font = Font(bold=True, size=12)
        year_label_cell.alignment = Alignment(horizontal='left')

        cover_ws.merge_cells(start_row=5, start_column=13, end_row=5, end_column=14)  # Modified here
        year_label_cell = cover_ws.cell(row=5, column=13, value=current_year)
        year_label_cell.font = Font(bold=True, size=10)

        cover_ws.merge_cells(start_row=5, start_column=3, end_row=5, end_column=6)  # Modified here
        value_year_cell = cover_ws.cell(row=5, column=3, value=current_month_name)
        value_year_cell.font = Font(bold=True, size=12)
        value_year_cell.border = border_style

        name_label_cell = cover_ws.cell(row=7, column=1, value="Name:")
        name_label_cell.font = Font(bold=True, size=12)

        cover_ws.merge_cells(start_row=7, start_column=3, end_row=7, end_column=6)  # Modified here
        name_label_cell = cover_ws.cell(row=7, column=3, value=str(user.username))
        name_label_cell.font = Font(bold=True, size=12)

        for column in range(3, 7):  # Columns 3 to 6 inclusive
            cell = cover_ws.cell(row=5, column=column)
            cell.border = border_style
        
        for column in range(3, 11):  # Columns 3 to 10 inclusive
            cell = cover_ws.cell(row=7, column=column)
            cell.border = border_style

        for column in range(13, 15):  # Columns 3 to 10 inclusive
            cell = cover_ws.cell(row=5, column=column)
            cell.border = border_style

        user_grade = user.get_grade()
        name_label_cell = cover_ws.cell(row=9, column=1, value="Category:")
        name_label_cell.font = Font(bold=True, size=12)

        cover_ws.cell(row=9, column=4, value="A1")
        cover_ws.cell(row=9, column=7, value="A2")
        cover_ws.cell(row=9, column=10, value="A3")

        cover_ws.cell(row=11, column=4, value="B1")
        cover_ws.cell(row=11, column=7, value="B2")
        cover_ws.cell(row=11, column=10, value="B3")

        cover_ws.cell(row=13, column=4, value="B4")
        cover_ws.cell(row=13, column=7, value="B5")

        # Define the rows and columns
        rows = [9, 11, 13]
        columns = [4, 7, 10]

        # Iterate over each cell and set the value and alignment
        for row in rows:
            for col in columns:
                cell = cover_ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal='center', vertical='center')


        # Mapping of grades to corresponding columns
        grade_mapping = {
            'A1': 'C9',
            'A2': 'F9',
            'A3': 'I9',
            'B1': 'C11',
            'B2': 'F11',
            'B3': 'I11',
            'B4': 'C13',
            'B5': 'F13',
        }

        # Set the letter '/' in the corresponding cell based on the user's grade
        if user_grade in grade_mapping:
            grade_column = grade_mapping[user_grade]
            cell_address = grade_column
            cover_ws[cell_address].value = '✓'
            cover_ws[cell_address].alignment = Alignment(horizontal='center', vertical='center')

        # Set borders for various cells
        utilities.set_other_borders(cover_ws, 9, ['C', 'F', 'I'])
        utilities.set_other_borders(cover_ws, 11, ['C', 'F', 'I'])
        utilities.set_other_borders(cover_ws, 13, ['C', 'F'])

        cell = cover_ws.cell(row=15, column=1, value="Nationality:")
        cell.font = Font(bold=True, size=12)

        cell = cover_ws.cell(row=15, column=6, value="Expatriate")
        cell.font = Font(name='Times New Roman', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')

        cell = cover_ws.cell(row=15, column=10, value="Local")
        cell.font = Font(name='Times New Roman', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Get the user's nationality
        user_nationality = user.get_expert()

        # Mapping of nationalities to corresponding columns
        nationality_mapping = {
            'EXP': 'D',  # Column for Expatriate
            'LOC': 'H',  # Column for Local
        }

        # Set the letter '/' in the corresponding cell and center it
        if user_nationality in nationality_mapping:
            nationality_column = nationality_mapping[user_nationality]
            cell_address = f'{nationality_column}15'
            cover_ws[cell_address].value = '✓'
            cover_ws[cell_address].alignment = Alignment(horizontal='center', vertical='center')

        utilities.set_other_borders(cover_ws, 15, ['D', 'H'])

        cell = cover_ws.cell(row=17, column=1, value="Group Field:")
        cell.font = Font(bold=True, size=12)

        cover_ws.merge_cells(start_row=17, start_column=5, end_row=17, end_column=9)
        cover_ws.merge_cells(start_row=17, start_column=11, end_row=17, end_column=17)
        cover_ws.merge_cells(start_row=19, start_column=5, end_row=19, end_column=9)
        cover_ws.merge_cells(start_row=19, start_column=11, end_row=19, end_column=17)

        cell = cover_ws.cell(row=17, column=5, value="Management and SHQE")
        cell.font = Font(size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')

        cell = cover_ws.cell(row=17, column=11, value="Tender Evaluation and Contract Negotiation")
        cell.font = Font(size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')

        cell = cover_ws.cell(row=19, column=5, value="Construction Supervision")
        cell.font = Font(size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')

        cell = cover_ws.cell(row=19, column=11, value="O&M")
        cell.font = Font(size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')

        utilities.set_other_borders(cover_ws, 17, ['D', 'J'])
        utilities.set_other_borders(cover_ws, 19, ['D', 'J'])

        grey_fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
        first_day_of_month = datetime(current_year, current_month, 1)
        shift_amount = 1
        for i in range(1, 17):
            col_address = chr(ord('A') + (i - 1 + shift_amount))  # Alternating columns C and G
            # Set the day of the week
            day_of_week_cell = cover_ws[col_address + '24']
            day_of_week_cell.value = (first_day_of_month + timedelta(days=i - 1)).strftime("%a")
            day_of_week_cell.alignment = Alignment(horizontal='center', vertical='center')
            day_of_week_cell.fill = grey_fill
            utilities.set_borders(cover_ws, 24, [col_address])
            utilities.set_borders(cover_ws, 26, [col_address])
            utilities.set_borders(cover_ws, 27, [col_address])
            utilities.set_borders(cover_ws, 28, [col_address])
            utilities.set_borders(cover_ws, 29, [col_address])
            # Set the day of the month
            day_of_month_cell = cover_ws[col_address + '25']
            day_of_month_cell.value = str((first_day_of_month + timedelta(days=i - 1)).day)
            day_of_month_cell.alignment = Alignment(horizontal='center', vertical='center')
            day_of_month_cell.fill = grey_fill
            utilities.set_borders(cover_ws, 25, [col_address])
            cell_address_activity_type = f'{col_address}26'

            # Replace 'activity_date' with the actual date for which you want to retrieve the activityType
            activity_date = first_day_of_month + timedelta(days=i - 1)
            activity_instance = Activity.objects.filter(user=user, activityDate=activity_date).first()

            if activity_instance:
                activity_type = activity_instance.get_activity_type()
                cover_ws[cell_address_activity_type].value = str(activity_type)
                if activity_type == constants.ACTIVITY_TYPES_CHOICES[constants.HOLIDAY][1]:
                    cover_ws[cell_address_activity_type].fill = grey_fill
            else:
                cover_ws[cell_address_activity_type].value = None
            cover_ws[cell_address_activity_type].alignment = Alignment(horizontal='center', vertical='center')

        first_day_of_second_half = datetime(current_year, current_month, 17).date()

        # Write abbreviated weekdays in row 30 for the second half of the month
        for i in range(0, calendar.monthrange(current_year, current_month)[1] - 16):
            col_address = chr(ord('A') + i + shift_amount) + '27'
            day_of_week = (first_day_of_second_half + timedelta(days=i)).strftime("%a")  # Get the abbreviated day name
            cell = cover_ws[col_address]
            cell.value = day_of_week
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = grey_fill
            utilities.set_borders(cover_ws, 27, [col_address])
            cell.border = Border(top=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'))

        # Continue writing the numeric values for the remaining days in the row below (row 30)
        for i in range(17, calendar.monthrange(current_year, current_month)[1] + 1):
            utilities.set_borders(cover_ws, 26, [col_address])
            col_address = chr(ord('A') + (i - 17 + shift_amount))
            cell_address = f'{col_address}28'
            cell = cover_ws[cell_address]
            cell.value = (first_day_of_month + timedelta(days=i - 1)).day
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = grey_fill
            utilities.set_borders(cover_ws, 28, [col_address])
            utilities.set_borders(cover_ws, 29, [col_address])

        for i in range(17, calendar.monthrange(current_year, current_month)[1] + 1):
            col_address_activity_type = chr(ord('A') + (i - 17 + shift_amount)) + '29'

            activity_date = datetime(current_year, current_month, 17) + timedelta(days=i - 17)
            activity_instance = Activity.objects.filter(user=user, activityDate=activity_date).first()

            cell = cover_ws[col_address_activity_type]
            if activity_instance:
                activity_type = activity_instance.get_activity_type()
                cell.value = str(activity_type)
                if activity_type == constants.ACTIVITY_TYPES_CHOICES[constants.HOLIDAY][1]:
                    cell.fill = grey_fill
            else:
                cell.value = None
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row in range(33, 36):
            for col_letter in ['D', 'F', 'H', 'J', 'L', 'N']:
                cell = cover_ws[col_letter + str(row)]
                cell.border = Border(
                    left=Side(style='thin', color='000000' if col_letter != 'C' else '000000'),
                    right=Side(style='thin', color='000000' if col_letter != 'J' else '000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )

        # add summary for working days
        cover_ws.merge_cells('D32:G33')
        cover_ws['D32'].value = "No. of Days (NOD)*"
        cover_ws['D32'].font = Font(size=11, bold=True)
        cover_ws['D32'].alignment = Alignment(horizontal='center', vertical='center')  # Center the text
        cover_ws['D32'].fill = grey_fill

        # Assuming user is an instance of the User model
        if user.expert == constants.LOCAL_USER:
            self.__add_local_working_days_individual__(current_date, user, cover_ws)
            # __add_local_working_days__(current_date, user, cover_ws)
        elif user.expert == constants.EXPERT_USER:
            self.__add_expert_working_days_individual__(current_date, user, cover_ws)

        # Merge cells and set values for total calendar days (TCD)
        cover_ws.merge_cells('H32:K33')
        tcd_cell = cover_ws['H32']
        tcd_cell.value = "Total Calendar Days\n(TCD)"
        tcd_cell.font = Font(size=10, bold=True)
        tcd_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        tcd_cell.fill = grey_fill

        # Set values for Japan and Cairo columns
        cover_ws.merge_cells('D34:E34')
        cover_ws.merge_cells('D35:E35')
        cover_ws.merge_cells('F34:G34')
        cover_ws.merge_cells('F35:G35')
        cover_ws.merge_cells('H34:I34')
        cover_ws.merge_cells('H35:I35')
        cover_ws.merge_cells('J34:K34')
        cover_ws.merge_cells('J35:K35')

        # Set values for Japan and Cairo columns
        labels = {
            'D34': 'Japan',
            'F34': 'Cairo',
            'H34': 'Japan',
            'J34': 'Cairo'
        }
        for cell_address, label_text in labels.items():
            cell = cover_ws[cell_address]
            cell.value = label_text
            cell.font = Font(size=11, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Merge cells and set values for consumption NOD/TCD
        cover_ws.merge_cells('L32:O33')
        consumption_cell = cover_ws['L32']
        consumption_cell.value = "Consumption NOD/TCD"
        consumption_cell.font = Font(size=11, bold=True)
        consumption_cell.alignment = Alignment(horizontal='center', vertical='center',  wrap_text=True)
        consumption_cell.fill = grey_fill

        cover_ws.merge_cells('L34:M34')
        cover_ws.merge_cells('N34:O34')
        cover_ws.merge_cells('L35:M35')
        cover_ws.merge_cells('N35:O35')

        # Set values for Japan and Cairo columns in consumption NOD/TCD section
        labels = {
            'L34': 'Japan',
            'N34': 'Cairo'
        }

        for cell_address, label_text in labels.items():
            cell = cover_ws[cell_address]
            cell.value = label_text
            cell.font = Font(size=11, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        cover_ws['G33'].border = Border(right=Side(style='thin', color='000000'))
        cover_ws['G32'].border = Border(right=Side(style='thin', color='000000'))
        cover_ws['D33'].border = Border(left=Side(style='thin', color='000000'))
        cover_ws['D32'].border = Border(left=Side(style='thin', color='000000'))

        cover_ws['K33'].border = Border(right=Side(style='thin', color='000000'))
        cover_ws['K32'].border = Border(right=Side(style='thin', color='000000'))
        cover_ws['H33'].border = Border(left=Side(style='thin', color='000000'))
        cover_ws['H32'].border = Border(left=Side(style='thin', color='000000'))

        cover_ws['O33'].border = Border(right=Side(style='thin', color='000000'))
        cover_ws['O32'].border = Border(right=Side(style='thin', color='000000'))
        cover_ws['L33'].border = Border(left=Side(style='thin', color='000000'))
        cover_ws['L32'].border = Border(left=Side(style='thin', color='000000'))

        for char in ['D', 'E', 'F', 'G', "H", "I", "J", "K", "L", "M", "N", "O"]:
            cover_ws[f'{char}32'].border = Border(top=Side(style='thin', color='000000'), left=Side(style='thin', color='000000'), right=Side(style="thin", color="000000"))

        # Project Director cell (B42)
        utilities.__format_cell__(cover_ws['C38'], "NOCE Approval")

        # NAT Approval cell (L42)
        utilities. __format_cell__(cover_ws['L38'], "NAT Approval")

        cover_ws.merge_cells(start_row=38, start_column=3, end_row=38, end_column=6)
        cover_ws.merge_cells(start_row=38, start_column=12, end_row=38, end_column=15)

        for col_letter in range(ord('A'), ord('S')):
            col_letter = chr(col_letter)
            cell = cover_ws[col_letter + '42']
            cell.border = Border(bottom=Side(style='thick'))
        
        for column in range(3, 7):  # Columns 3 to 10 inclusive
            cell = cover_ws.cell(row=41, column=column)
            cell.border = border_style
        
        for column in range(12, 16):  # Columns 3 to 10 inclusive
            cell = cover_ws.cell(row=41, column=column)
            cell.border = border_style

        cover_ws.merge_cells("B43:F43")
        cover_ws.merge_cells("M43:Q43")
        labels = {
            'B43': "J = Working day In Japan",
            'M43': "C = Working day In Cairo",
            'B44': "H = Official Holiday In Cairo",
            'M44': "X = Day off",
            'B46': "Note: According to the contract 81/M the total days are working days in Cairo plus to official holiday in Egypt\n *NOD=C (Working day in Cairo)+H (Official Holiday in Egypt)"
        }

        cover_ws.merge_cells(start_row=46, start_column=2, end_row=48, end_column=17)
        cell = cover_ws['B46']
        cell.value = labels['B46']

        # Enable text wrapping
        cell.alignment = Alignment(wrap_text=True)

        for cell_address, label_text in labels.items():
            cover_ws.merge_cells(f'{cell_address}:{chr(ord(cell_address[0]) + 2)}{cell_address[1:]}')
            cell = cover_ws[cell_address]
            cell.value = label_text
            cell.font = Font(size=11, bold=True)

        # Set column width to 4.5
        for col_letter in range(ord('A'), ord('S')):
            col_letter = chr(col_letter)
            cover_ws.column_dimensions[col_letter].width = 4.5
        
        for row in range(48, cover_ws.max_row + 1):
            cover_ws.row_dimensions[row].hidden = True
        
        # Hide columns beyond S
        for col in range(19, cover_ws.max_column + 1):
            cover_ws.column_dimensions[get_column_letter(col)].hidden = True
        
        for row in cover_ws.iter_rows():
            for cell in row:
                cover_ws.row_dimensions[cell.row].height = 15

            
    @action(detail=False, methods=['GET'], url_path=r'extract_timesheet/(?P<userId>\w+(?:-\w+)*)', serializer_class=UserTimeSheetSerializer)
    def extract_timesheet(self, request, *args, **kwargs):
        if not utilities.check_is_admin(request.user.id):
            return Response(status=status.HTTP_403_FORBIDDEN)

        serializer = UserTimeSheetSerializer(data=request.query_params)
        serializer.is_valid()

        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        # get the month and user id
        date = serializer.validated_data['date']
        user_id = kwargs['userId']

        # get all the activities for that month by that user
        activities = Activity.objects.filter(activityDate__month=date.month, activityDate__year=date.year, user__id=user_id).order_by('-created')
        user = User.objects.get(id=user_id)

        dateFont = Font(size=16)
        wb = Workbook()

        ws = wb.active
        ws.title = "User Timesheet"

        month = date.month
        year = date.year

        name_year_month_border = Border(
            left=Side(border_style='thin'),
            right=Side(border_style='thin'),
            top=Side(border_style='thin'),
            bottom=Side(border_style='thin')
        )

        # Merge cells for the new section
        ws.merge_cells(start_row=2, start_column=2, end_row=5, end_column=10)
        for row in range(2, 6):
            for col in range(2, 11):
                cell = ws.cell(row=row, column=col)
                cell.border = name_year_month_border
        
        # Set the text and alignment in the first cell of the merged range
        merged_cell = ws.cell(row=2, column=2)
        merged_cell.value = "Consulting Services for Greater Cairo Metro Line No. 4 Phase1 Project"
        merged_cell.font = dateFont
        merged_cell.alignment = Alignment(horizontal='center', vertical='center')

        script_directory = os.path.dirname(os.path.abspath(__file__))
        parent_parent_directory = os.path.dirname(script_directory)
        logo_path = os.path.join(parent_parent_directory, 'static', 'images', 'logo.png')
        img = Image(logo_path)
        img.height = 1.08 * 72  # 1 inch = 72 points
        img.width = 1.14 * 72
        ws.add_image(img, 'M2')

        # Set font and border for "Year:" label in cell A8
        year_label_cell = ws.cell(row=8, column=1, value="Year:")
        year_label_cell.font = Font(bold=True, size=12)
        year_label_cell.border = name_year_month_border

        # Align "Year:" label horizontally and vertically
        year_label_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Set font and border for the year value in cell B8
        year_value_cell = ws.cell(row=8, column=2, value=year)
        year_value_cell.font = Font(bold=True, size=12)
        year_value_cell.border = name_year_month_border

        # Align year value horizontally and vertically
        year_value_cell.alignment = Alignment(horizontal='center', vertical='center')
        year_digits = year % 100
        month_name = calendar.month_name[month]

        # Combine the month name and the last two digits of the year
        formatted_month = f"{month_name}-{year_digits:02d}"

        cell = ws.cell(row=9, column=2, value=formatted_month)
        cell.border = name_year_month_border
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')  # Center the text horizontally and vertically

        # Set the text "Month" in cell A9
        cell_month = ws.cell(row=9, column=1, value="Month")
        cell_month.border = name_year_month_border
        cell_month.font = Font(bold=True, size=12)
        cell_month.alignment = Alignment(horizontal='center', vertical='center')  # Center the text horizontally and vertically

        # Adjust the width of column A to fit the text
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 10 # Adjust the height as needed

        ws.merge_cells(start_row=8, start_column=7, end_row=8, end_column=11)
        ws.merge_cells(start_row=9, start_column=7, end_row=9, end_column=11)
        ws.merge_cells(start_row=12, start_column=5, end_row=12, end_column=14)

        for row in range(12, 13):
            for col in range(6, 15):
                cell = ws.cell(row=row, column=col)
                cell.border = name_year_month_border

        cell = ws.cell(row=8, column=7, value=user.username)
        cell.font = Font(bold=True, size=12)

        # Set alignment
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Create border
        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))

        # Apply border to each cell in the merged range
        for row in ws.iter_rows(min_row=8, max_row=8, min_col=7, max_col=11):
            for cell in row:
                cell.border = border

        cell = ws.cell(row=9, column=7, value=user.position)
        cell.font = Font(bold=True, size=12)

        # Set alignment
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Create border
        border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))

        # Apply border to each cell in the merged range
        for row in ws.iter_rows(min_row=9, max_row=9, min_col=7, max_col=11):
            for cell in row:
                cell.border = border

        # Create headers for columns
        headers_border = Border(left=Side(style='medium'), 
                right=Side(style='medium'), 
                top=Side(style='medium'), 
                bottom=Side(style='medium'))
        
        for col in range(1, 4):
            header_cell = ws.cell(row=12, column=col, value=["Day", "Cairo", "Japan"][col - 1])
            header_cell.font = dateFont
            header_cell.border = headers_border
            header_cell.alignment = Alignment(textRotation=90, vertical='center')

        cell = ws.cell(row=12, column=5, value="DAILY ACTIVITIES")
        cell.font = dateFont
        cell.border = name_year_month_border

        # Center the text horizontally and vertically
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for day in range(1, calendar.monthrange(year, month)[1] + 1):
            start_row_index = day + 12  # Offset by 12 to account for existing rows]

            # Set value for the first cell in the merged range (representing the day)
            cell = ws.cell(row=start_row_index, column=1)
            cell.value = f"{day:02d}"
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Set borders for each cell within the merged range
            for col in ['A', 'B', 'C']:
                for row in range(start_row_index, start_row_index + 1):
                    cell_address = f'{col}{row}'
                    ws[cell_address].border = Border(top=Side(style='thin', color='000000'),
                                                                    left=Side(style='thin', color='000000'),
                                                                    right=Side(style='thin', color='000000'),
                                                                    bottom=Side(style='thin', color='000000'))
            
            # Fetch activities for the current day
            activities_for_day = activities.filter(activityDate__day=day)

            # Combine activities' text and type
            activities_text = "\n".join([activity.userActivity or '' for activity in activities_for_day])
            activities_type = "\n".join([str(activity.get_activity_type()) for activity in activities_for_day])

            # Calculate the number of merged rows for activities
            num_merged_rows = activities_text.count('\n') + 1

            for col in range(5, 14):
                bottom_cell_address = f"{get_column_letter(col)}{start_row_index + num_merged_rows - 1}"
                ws[bottom_cell_address].border = Border(bottom=Side(style='thin', color='000000'))

            merged_range = f"E{start_row_index}:N{start_row_index}"  # Update the range accordingly
            ws.merge_cells(merged_range)
            merged_cell = ws.cell(row=start_row_index, column=5)  # Top-left cell of the merged range
            merged_cell.value = activities_text
            merged_cell.font = Font(size=10)
            merged_cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
        
            font_size = 10
            line_height = 1.2 * font_size
            row_height_factor = 2
            required_height = line_height * num_merged_rows * row_height_factor

            # Adjust the row height for each merged row
            for i in range(start_row_index, start_row_index + num_merged_rows):
                row_dimension = ws.row_dimensions[i]
                row_dimension.height = required_height

            # Set the activities type in the appropriate column
            if user.expert in [constants.LOCAL_USER, constants.EXPERT_USER]:
                cell = ws.cell(row=start_row_index, column=3 if activities_type == "J" else 2)
                cell.value = activities_type
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(size=11)

        # Get the current month and year
        current_month_name = date.strftime("%B")

        self.__add_individual_timesheet__(wb=wb, current_month_name=current_month_name,\
                                           current_year=date.year, user=user, current_date=date,\
                                            current_month=date.month)

        excel_data = BytesIO()
        wb.save(excel_data)
        excel_data.seek(0)
        response = HttpResponse(excel_data, content_type="application/ms-excel")
        response["Content-Disposition"] = f'attachment; filename=timesheet{date.year}_{date.month}_{user_id}.xlsx'
        return response

class ActivityViewSet(viewsets.ModelViewSet):
    pagination_class = PageNumberPagination
    permission_classes=(IsAuthenticated,)
    queryset = Activity.objects.all()

    @action(detail=False, methods=['GET'])
    def link(self, request, *args, **kwargs):
        allActivities = Activity.objects.all()
        data = []
        for activity in allActivities:
            if activity.user is None:
                user_details = activity.user_details
                if user_details:
                    user = User.objects.filter(email=user_details['email']).first()
                    if user:
                        activity.user = user
                        activity.save()
                        data.append(user)
        serializer = ListUsersSerializer(data, many=True)
        return Response(data=serializer.data, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['GET'])
    def deleted_user_activities(self, request, *args, **kwargs):
        allActivities = Activity.objects.all()
        data = []
        for activity in allActivities:
            if activity.user is None:
                data.append(activity.user)
        serializer = ListUsersSerializer(data, many=True)
        return Response(data=serializer.data, status=status.HTTP_200_OK)

    def get_serializer_class(self):
        return ActivitySerializer

    @action(detail=False, methods=['GET'])
    def get_activities(self, request, *args, **kwargs):
        requestId = request.user.id
        user = User.objects.filter(id=requestId).first()
        date_param = request.query_params.get('date', None)

        if date_param:
        # If a date is provided, parse it and fetch activities for that date
            try:
                selected_date = datetime.strptime(date_param, '%Y-%m-%d').date()
            except ValueError:
                return Response({'detail': 'Invalid date format. Please use YYYY-MM-DD format.'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            # If no date is provided, default to today's date
            selected_date = date.today()
        if user.is_superuser or user.isAdmin:
            activities = Activity.objects.all().order_by('-created')
        else:
            activities = Activity.objects.filter(user_id=requestId).order_by('-created')
        serializer = ActivitySerializer(activities, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['GET'])
    def my_activities(self, request, *args, **kwargs):
        """
            This returns the logged in user activities for the current month,
            however if the logged in user is a superuser/admin then it shows
            all users' activities
        """
        requestId = request.user.id
        user = User.objects.filter(id=requestId).first()

        serializer = UserTimeSheetSerializer(data=request.query_params)
        serializer.is_valid()
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        # Get the first day and last day of the current month
        current_date = serializer.validated_data['date']

        first_day_of_month = current_date.replace(day=1)
        last_day_of_month = (first_day_of_month + timezone.timedelta(days=32)).replace(day=1) - timezone.timedelta(days=1)

        if (user.is_superuser or user.isAdmin):
            activities = Activity.objects.filter(activityDate__range=[first_day_of_month, last_day_of_month]).order_by('activityDate')
        else:
            # Filter activities for the current month and order by activityDate
            activities = Activity.objects.filter(user=user, activityDate__range=[first_day_of_month, last_day_of_month]).order_by('activityDate')
        # Paginate the queryset
        page = self.paginate_queryset(activities)
        if page is not None:
            serializer = ActivitySerializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = ActivitySerializer(activities, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def __format_cell__(self, cell, value, size=11, italic=False, center=True):
        cell.value = value
        cell.font = Font(size=size, italic=True)
        if center:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def __add_local_working_days__(self, current_date, user, cover_ws):
        ##### EXPERT #####
        start_date = current_date.date().replace(day=1)
        last_day_of_month = calendar.monthrange(current_date.date().year, current_date.date().month)[1]
        end_date = (current_date.date() + timedelta(days=last_day_of_month)).replace(day=1)

        # Adjust end_date to the next day
        end_date = end_date + timedelta(days=1)

        total_working_days_expert = np.busday_count(start_date, end_date, weekmask='0011111')
        cover_ws['F42'].value = total_working_days_expert
        cover_ws['F42'].alignment = Alignment(horizontal='center', vertical='center')  # Center the text
        ##### EXPERT #####

        start_date = current_date.date().replace(day=1)
        last_day_of_month = calendar.monthrange(current_date.date().year, current_date.date().month)[1]
        end_date = current_date.date().replace(day=last_day_of_month)

        last_day_of_month = calendar.monthrange(current_date.date().year, current_date.date().month)[1]
        end_date = current_date.date().replace(day=last_day_of_month) + timedelta(days=1)
        total_working_days = np.busday_count(start_date, end_date, weekmask='1111111')
        # Iterate through each day in the range
        for day in range(1, last_day_of_month + 1):
            day_off = Activity.objects.filter(
                user=user,
                activityDate__day=current_date.date().day,
                activityDate__month=current_date.date().month,
                activityDate__year=current_date.date().year,
                activityType=constants.OFFDAY,
            ).exists()

            if day_off and (day == 5 and day - 1 in (4, 6) and day + 1 in (4, 6)):
                total_working_days -= 1

            # Filter activities for the current user and month excluding 'H' type activities
            activities = Activity.objects.filter(
                user=user,
                activityDate__month=current_date.date().month,
                activityDate__year=current_date.date().year,
            ).exclude(activityType=constants.OFFDAY)

            # Count the number of activities
            working_days = activities.count()
            # Iterate over weeks in the month
            for week in calendar.monthcalendar(current_date.year, current_date.month):
                for day in week:
                    # Check if Thursday and Saturday are off-days in the current week
                    if day != 0:  # Ignore days that belong to the previous or next month (represented as 0)
                        thursday_offday = any(activity.activityType == constants.OFFDAY for activity in activities.filter(activityDate__day=day + 3))
                        saturday_offday = any(activity.activityType == constants.OFFDAY for activity in activities.filter(activityDate__day=day + 5))

                        # If either Thursday or Saturday is off-day, decrement working_days
                        if thursday_offday or saturday_offday:
                            working_days -= 1

        # Add "0" under "Cairo" for local users
        cover_ws['D42'].value = working_days
        cover_ws['D42'].alignment = Alignment(horizontal='center', vertical='center')  # Center the text
        cover_ws['G42'].value = total_working_days
        cover_ws['G42'].alignment = Alignment(horizontal='center', vertical='center')  # Center the text

        cover_ws['C42'].value = 0
        cover_ws['C42'].alignment = Alignment(horizontal='center', vertical='center')  # Center the text

        cover_ws['J42'].value = round(cover_ws['D42'].value / cover_ws['G42'].value, 3)
        cover_ws['J42'].font = Font(size=11)
        cover_ws['J42'].alignment = Alignment(horizontal='center', vertical='center')

         # Japan NOD/TCD
        cover_ws['I42'].value = round(cover_ws['C42'].value / cover_ws['F42'].value, 3)
        cover_ws['I42'].font = Font(size=11)
        cover_ws['I42'].alignment = Alignment(horizontal='center', vertical='center')

    def __add_expert_working_days__(self, current_date, user, cover_ws):
        ###### LOCAL ######
        start_date = current_date.date().replace(day=1)
        last_day_of_month = calendar.monthrange(current_date.date().year, current_date.date().month)[1]
        end_date = current_date.date().replace(day=last_day_of_month)

        last_day_of_month = calendar.monthrange(current_date.date().year, current_date.date().month)[1]
        end_date = current_date.date().replace(day=last_day_of_month) + timedelta(days=1)
        total_working_days_cairo = np.busday_count(start_date, end_date, weekmask='1111111')
         ###### LOCAL ######
        start_date = current_date.date().replace(day=1)
        last_day_of_month = calendar.monthrange(current_date.date().year, current_date.date().month)[1]
        end_date = (current_date.date() + timedelta(days=last_day_of_month)).replace(day=1)

        # Adjust end_date to the next day
        end_date = end_date + timedelta(days=1)
        total_working_days = np.busday_count(start_date, end_date, weekmask='0011111')

        # Filter activities for the current user and month excluding 'H' type activities
        activities = Activity.objects.filter(
            user=user,
            activityDate__month=current_date.date().month,
            activityDate__year=current_date.date().year,
        ).exclude(activityType=constants.OFFDAY)

        # Count the number of activities
        working_days = activities.count()

        cover_ws['C42'].value = working_days
        cover_ws['C42'].alignment = Alignment(horizontal='center', vertical='center')  # Center the text

        cover_ws['F42'].value = total_working_days
        cover_ws['F42'].alignment = Alignment(horizontal='center', vertical='center')  # Center the text

        # Cairo NOD
        cover_ws['D42'].value = 0
        cover_ws['D42'].alignment = Alignment(horizontal='center', vertical='center')  # Center the text

        # Cairo TCD
        cover_ws['G42'].value = total_working_days_cairo
        cover_ws['G42'].alignment = Alignment(horizontal='center', vertical='center')  # Center the text

        # Japan NOD/TCD
        cover_ws['I42'].value = round(cover_ws['C38'].value / cover_ws['F42'].value, 3)
        cover_ws['I42'].font = Font(size=11)
        cover_ws['I42'].alignment = Alignment(horizontal='center', vertical='center')

        # Cairo NOD/TCD
        cover_ws['J42'].value = 0
        cover_ws['J42'].font = Font(size=11)
        cover_ws['J42'].alignment = Alignment(horizontal='center', vertical='center')

    def set_borders(self, ws, row, columns):
        for col in columns:
            cell_address = f'{col}{row}'
            ws[cell_address].border = Border(top=Side(style='thin', color='000000'),
                                            left=Side(style='thin', color='000000'),
                                            right=Side(style='thin', color='000000'),
                                            bottom=Side(style='thin', color='000000'))

    @action(detail=False, methods=['GET'])
    def export_all(self, request, *args, **kwargs):
        user = request.user
        if not (user.is_superuser or user.isAdmin):
            return Response(constants.NOT_ALLOWED_TO_ACCESS, status=status.HTTP_400_BAD_REQUEST)
        
        serializer = UserTimeSheetSerializer(data=request.query_params)
        serializer.is_valid()

        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        date = serializer.validated_data.get('date')

        company_param = request.query_params.get('company', None)
        if company_param:
            company_users = User.objects.filter(company=company_param)
            return generate_noce_timesheet(users=company_users, companyName=constants.COMPANY_CHOICES[int(company_param)][1], date=date)
        else:
            return generate_noce_timesheet(date=date)

        # return Response(status=status.HTTP_200_OK)

    @action(detail=False, methods=['POST'])
    def create_activity(self, request, *args, **kwargs):
        """
            This endpoint allows the user to create an activity
        """
        userId = request.user.id

        if User.objects.filter(id=userId).first().needsPasswordReset:
            return Response(constants.ERR_PASSWORD_RESET_NEEDED, status=status.HTTP_400_BAD_REQUEST)

        serializer = CreateActivitySerializer(data=request.data)
        serializer.is_valid()
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        userActivity = serializer.validated_data.get('userActivity', None)
        activityType = serializer.validated_data.get('activityType', None)
        activityDate = serializer.validated_data.get('activityDate', None)
        # Check if the user has already logged an activity today
        existing_activity = Activity.objects.filter(user__id=userId, activityDate=activityDate).first()

        if existing_activity:
            return Response({"detail": "You have already logged an activity for the selected date."}, status=status.HTTP_400_BAD_REQUEST)

        new_activity = Activity.objects.create(userActivity=userActivity,\
                                                activityType=activityType,\
                                                user_id=userId,
                                                activityDate=activityDate)
        new_activity.save()
        return Response(constants.SUCCESSFULLY_CREATED_ACTIVITY, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['DELETE'], url_path=r'delete_activity/(?P<activityId>\w+(?:-\w+)*)')
    def delete_activity(self, request, *args, **kwargs):
        """
            This endpoint allows the user/admin to delete an activity
        """
        userId = request.user.id

        activity = Activity.objects.filter(id=kwargs['activityId']).first()
        user = User.objects.filter(id=userId).first()

        if not activity:
            return Response(status=status.HTTP_404_NOT_FOUND)

        if userId == activity.user.id or user.is_superuser:
            activity.delete()
            return Response(constants.SUCCESSFULLY_DELETED_ACTIVITY, status=status.HTTP_200_OK)

        return Response(status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['GET'])
    def calculate_activity(self, request, *args, **kwargs):
        """
        This endpoint calculates how many activities a user does out of the possible working days
        EXPERT -> Saturday, Sunday off
        LOCAL -> Friday off
        """
        userId = request.user.id
        if not utilities.check_is_admin(userId):
            return Response(status=status.HTTP_403_FORBIDDEN)

        serializer = UserTimeSheetSerializer(data=request.query_params)
        serializer.is_valid()

        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        date = serializer.validated_data['date']

        # Get all users
        users = User.objects.all()

        data = []

        for user in users:
            start_date = date.replace(day=1)
            last_day_of_month = calendar.monthrange(date.year, date.month)[1]
            end_date = date.replace(day=last_day_of_month)

            if user.expert == constants.EXPERT_USER:
                start_date = date.replace(day=1)
                end_date = date.replace(day=calendar.monthrange(date.year, date.month)[1])
                end_date = end_date + relativedelta(days=1)

                total_working_days_japan = constants.JAPAN_WORKING_DAYS[start_date.month]
                # Filter activities for the current user and month excluding 'H' type activities
                activities = Activity.objects.filter(
                    user=user,
                    activityDate__month=date.month,
                    activityDate__year=date.year,
                    activityType__in=[constants.HOMEASSIGN]
                )

                # Count the number of activities
                working_days_japan = activities.count()
                last_day_of_month = calendar.monthrange(date.year, date.month)[1]
                end_date = date.replace(day=last_day_of_month) + timedelta(days=1)

                total_working_days_cairo = np.busday_count(start_date, end_date, weekmask='1111111')
                working_days_cairo = Activity.objects.filter(
                    user=user,
                    activityDate__month=date.month,
                    activityDate__year=date.year,
                ).filter(activityType__in=[constants.HOLIDAY, constants.INCAIRO]).count()
                data.append({
                    'user__id': user.id,
                    'user__first_name': user.first_name,
                    'user__last_name': user.last_name,
                    'user__email': user.email,
                    'working_days_cairo': working_days_cairo,
                    'total_days_cairo': total_working_days_cairo,
                    'working_days_japan': working_days_japan,
                    'total_days_japan': total_working_days_japan
                })

            if user.expert == constants.LOCAL_USER:
                # Get the last day of the month
                last_day_of_month = calendar.monthrange(date.year, date.month)[1]
                end_date = date.replace(day=last_day_of_month) + timedelta(days=1)
                total_working_days_cairo = np.busday_count(start_date, end_date, weekmask='1111111')
                total_working_days_japan = constants.JAPAN_WORKING_DAYS[start_date.month]
                # Iterate through each day in the range

                # Filter activities for the current user and month excluding 'H' type activities
                activities_cairo = Activity.objects.filter(
                    user=user,
                    activityDate__month=date.month,
                    activityDate__year=date.year,
                    activityType__in=[constants.HOLIDAY, constants.INCAIRO]
                )

                # Count the number of activities
                working_days_cairo = activities_cairo.count()

                all_cairo_activities = Activity.objects.filter(
                    user=user,
                    activityDate__month=date.month,
                    activityDate__year=date.year,
                    activityType__in=[constants.HOLIDAY, constants.INCAIRO]
                )

                for week in calendar.monthcalendar(date.year, date.month):
                    for day in week:
                        # Check if Thursday and Saturday are off-days in the current week
                        if day != 0:  # Ignore days that belong to the previous or next month (represented as 0)
                            thursday_offday = any(activity.activityType == constants.OFFDAY for activity in all_cairo_activities.filter(activityDate__day=day + 3))
                            friday_offday = any(activity.activityType == constants.OFFDAY for activity in all_cairo_activities.filter(activityDate__day=day + 4))
                            saturday_offday = any(activity.activityType == constants.OFFDAY for activity in all_cairo_activities.filter(activityDate__day=day + 5))

                            # If either Thursday or Saturday is off-day, decrement working_days
                            if thursday_offday and saturday_offday and friday_offday:
                                total_working_days_cairo -= 1

                start_date = date.replace(day=1)
                end_date = date.replace(day=calendar.monthrange(date.year, date.month)[1])
                end_date = end_date + relativedelta(days=1)

                total_working_days_japan = constants.JAPAN_WORKING_DAYS[start_date.month]
                np.busday_count(start_date, end_date, weekmask='0011111')

                activities_japan = Activity.objects.filter(
                    user=user,
                    activityDate__month=date.month,
                    activityDate__year=date.year,
                ).filter(activityType__in=[constants.HOMEASSIGN])
                working_days_japan = activities_japan.count()

                # Append user data to the response
                data.append({
                    'user__id': user.id,
                    'user__first_name': user.first_name,
                    'user__last_name': user.last_name,
                    'user__email': user.email,
                    'working_days_cairo': working_days_cairo,
                    'total_days_cairo': total_working_days_cairo,
                    'working_days_japan': working_days_japan,
                    'total_days_japan': total_working_days_japan
                })

        return Response(data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['PATCH'], url_path=r'edit_activity/(?P<activityId>\w+(?:-\w+)*)')
    def edit_activity(self, request, *args, **kwargs):
        serializer = EditActivitySerializer(data=request.data)
        serializer.is_valid()
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        request_user_id = request.user.id
        activityId = self.kwargs['activityId']

        # Check if activity exists
        activity = Activity.objects.filter(id=activityId).first()
        if not activity:
            return Response(constants.ERR_NO_ACTIVITY_ID_FOUND, status=status.HTTP_400_BAD_REQUEST)

        user = User.objects.filter(id=request_user_id).first()

        # check if user editing is the owner of that activity
        if activity.user.id != request_user_id and not (user.isAdmin or user.is_superuser):
            return Response(constants.NOT_ALLOWED_TO_ACCESS, status=status.HTTP_400_BAD_REQUEST)


        # Update the activity fields with the serializer data
        activity.userActivity = serializer.validated_data.get('userActivity', activity.userActivity)
        activity.activityType = serializer.validated_data.get('activityType', activity.activityType)
        activity.activityDate = serializer.validated_data.get('activityDate', activity.activityDate)
        activity.save()

        return Response(status=status.HTTP_200_OK)

    @action(detail=False, methods=['POST'], url_path=r'admin/log/(?P<userId>\w+(?:-\w+)*)')
    def admin_log_activity(self, request, *args, **kwargs):
        requestId = request.user.id
        if not utilities.check_is_admin(requestId):
            return Response(status=status.HTTP_403_FORBIDDEN)

        userId = kwargs['userId']

        serializer = CreateActivitySerializer(data=request.data)
        serializer.is_valid()
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        userActivity = serializer.validated_data.get('userActivity', None)
        activityType = serializer.validated_data.get('activityType', None)
        activityDate = serializer.validated_data.get('activityDate', None)
        # Check if the user has already logged an activity today
        existing_activity = Activity.objects.filter(user__id=userId, activityDate=activityDate).first()

        if existing_activity:
            return Response({"detail": "An activity for the selected date already exists. Please delete it firstly."}, status=status.HTTP_400_BAD_REQUEST)

        new_activity = Activity.objects.create(userActivity=userActivity,\
                                                activityType=activityType,\
                                                user_id=userId,
                                                activityDate=activityDate)
        new_activity.save()

        return Response(status=status.HTTP_201_CREATED)

class LatestFileView(viewsets.ModelViewSet):
    @action(detail=False, methods=['GET'], url_path=r'activities/own_timesheet/(?P<userId>\w+(?:-\w+)*)')
    def export_own_timesheet(self, request, *args, **kwargs):
        serializer = UserTimeSheetSerializer(data=request.query_params)
        serializer.is_valid()

        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        user_id = kwargs['userId']

        # Get the first day and last day of the current month
        date = serializer.validated_data['date']

        activities = Activity.objects.filter(activityDate__month=date.month, activityDate__year=date.year, user__id=user_id).order_by('-created')
        user = User.objects.get(id=user_id)

        dateFont = Font(size=16)
        wb = Workbook()

        ws = wb.active
        ws.title = "My Timesheet"

        month = date.month
        year = date.year

        name_year_month_border = Border(
            left=Side(border_style='thin'),
            right=Side(border_style='thin'),
            top=Side(border_style='thin'),
            bottom=Side(border_style='thin')
        )

        # Merge cells for the new section
        ws.merge_cells(start_row=2, start_column=2, end_row=5, end_column=10)
        for row in range(2, 6):
            for col in range(2, 11):
                cell = ws.cell(row=row, column=col)
                cell.border = name_year_month_border
        
        # Set the text and alignment in the first cell of the merged range
        merged_cell = ws.cell(row=2, column=2)
        merged_cell.value = "Consulting Services for Greater Cairo Metro Line No. 4 Phase1 Project"
        merged_cell.font = dateFont
        merged_cell.alignment = Alignment(horizontal='center', vertical='center')

        script_directory = os.path.dirname(os.path.abspath(__file__))
        parent_parent_directory = os.path.dirname(script_directory)
        logo_path = os.path.join(parent_parent_directory, 'static', 'images', 'logo.png')
        img = Image(logo_path)
        ws.add_image(img, 'M2')

        # Set font and border for "Year:" label in cell A8
        year_label_cell = ws.cell(row=8, column=1, value="Year:")
        year_label_cell.font = Font(bold=True, size=12)
        year_label_cell.border = name_year_month_border

        # Align "Year:" label horizontally and vertically
        year_label_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Set font and border for the year value in cell B8
        year_value_cell = ws.cell(row=8, column=2, value=year)
        year_value_cell.font = Font(bold=True, size=12)
        year_value_cell.border = name_year_month_border

        # Align year value horizontally and vertically
        year_value_cell.alignment = Alignment(horizontal='center', vertical='center')
        year_digits = year % 100
        month_name = calendar.month_name[month]

        # Combine the month name and the last two digits of the year
        formatted_month = f"{month_name}-{year_digits:02d}"

        cell = ws.cell(row=9, column=2, value=formatted_month)
        cell.border = name_year_month_border
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')  # Center the text horizontally and vertically

        # Set the text "Month" in cell A9
        cell_month = ws.cell(row=9, column=1, value="Month")
        cell_month.border = name_year_month_border
        cell_month.font = Font(bold=True, size=12)
        cell_month.alignment = Alignment(horizontal='center', vertical='center')  # Center the text horizontally and vertically

        # Adjust the width of column A to fit the text
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 10 # Adjust the height as needed

        ws.merge_cells(start_row=8, start_column=7, end_row=8, end_column=11)
        ws.merge_cells(start_row=9, start_column=7, end_row=9, end_column=11)
        ws.merge_cells(start_row=12, start_column=6, end_row=12, end_column=15)

        for row in range(12, 13):
            for col in range(6, 16):
                cell = ws.cell(row=row, column=col)
                cell.border = name_year_month_border

        cell = ws.cell(row=8, column=7, value=user.get_full_name())
        cell.font = Font(bold=True, size=12)

        # Set alignment
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Create border
        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))

        # Apply border to each cell in the merged range
        for row in ws.iter_rows(min_row=8, max_row=8, min_col=7, max_col=11):
            for cell in row:
                cell.border = border

        cell = ws.cell(row=9, column=7, value=user.position)
        cell.font = Font(bold=True, size=12)

        # Set alignment
        cell.alignment = Alignment(horizontal='center', vertical='center')

        # Create border
        border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))

        # Apply border to each cell in the merged range
        for row in ws.iter_rows(min_row=9, max_row=9, min_col=7, max_col=11):
            for cell in row:
                cell.border = border

        # Create headers for columns
        headers_border = Border(left=Side(style='medium'), 
                right=Side(style='medium'), 
                top=Side(style='medium'), 
                bottom=Side(style='medium'))
        
        for col in range(1, 4):
            header_cell = ws.cell(row=12, column=col, value=["Day", "Cairo", "Japan"][col - 1])
            header_cell.font = dateFont
            header_cell.border = headers_border
            header_cell.alignment = Alignment(textRotation=90, vertical='center')

        cell = ws.cell(row=12, column=6, value="DAILY ACTIVITIES")
        cell.font = dateFont
        cell.border = name_year_month_border

        # Center the text horizontally and vertically
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for day in range(1, calendar.monthrange(year, month)[1] + 1):
            start_row_index = 2 * day + 11  # Offset by 12 to account for existing rows]
            
            # Merge two rows for each day in columns 1, 2, and 3
            ws.merge_cells(start_row=start_row_index, start_column=1, end_row=start_row_index + 1, end_column=1)
            ws.merge_cells(start_row=start_row_index, start_column=2, end_row=start_row_index + 1, end_column=2)
            ws.merge_cells(start_row=start_row_index, start_column=3, end_row=start_row_index + 1, end_column=3)
            
            # Set value for the first cell in the merged range (representing the day)
            cell = ws.cell(row=start_row_index, column=1)
            cell.value = f"{day:02d}"
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Set borders for each cell within the merged range
            for col in ['A', 'B', 'C']:
                for row in range(start_row_index, start_row_index + 2):
                    cell_address = f'{col}{row}'
                    ws[cell_address].border = Border(top=Side(style='thin', color='000000'),
                                                                    left=Side(style='thin', color='000000'),
                                                                    right=Side(style='thin', color='000000'),
                                                                    bottom=Side(style='thin', color='000000'))
            
            # Fetch activities for the current day
            activities_for_day = activities.filter(activityDate__day=day)

            # Combine activities' text and type
            activities_text = "\n".join([activity.userActivity or '' for activity in activities_for_day])
            activities_type = "\n".join([str(activity.get_activity_type()) for activity in activities_for_day])

            # Calculate the number of merged rows for activities
            num_merged_rows = activities_text.count('\n') + 1

            # Merge cells for the "Activities" column (two rows)
            start_column_letter = get_column_letter(6)  # Column D
            end_column_letter = get_column_letter(15)  # Adjust the last column letter as needed
            activities_column_range = f"{start_column_letter}{start_row_index}:{end_column_letter}{(start_row_index + 2)-1}"
            ws.merge_cells(activities_column_range)

            for col in range(6, 16):
                bottom_cell_address = f"{get_column_letter(col)}{start_row_index + num_merged_rows}"
                ws[bottom_cell_address].border = Border(bottom=Side(style='thin', color='000000'))

            # Set value and font for the cell in column 6 (representing activities)
            
            activities_cell = ws.cell(row=start_row_index, column=6)
            activities_cell.value = activities_text
            activities_cell.font = Font(size=10)

            # Adjust row height to fit the content of the cell in column 6
            activities_cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center', indent=5)
            
            # Calculate the required row height based on the number of lines in the text
            font_size = 10
            line_height = 1.2 * font_size
            required_height = line_height * num_merged_rows

            # Adjust the row height for each merged row
            for i in range(start_row_index, start_row_index + num_merged_rows):
                row_dimension = ws.row_dimensions[i]
                row_dimension.height = required_height

            # Set the activities type in the appropriate column
            if user.expert in [constants.LOCAL_USER, constants.EXPERT_USER]:
                cell = ws.cell(row=start_row_index, column=3 if activities_type == "J" else 2)
                cell.value = activities_type
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(size=11)

        excel_data = BytesIO()
        wb.save(excel_data)
        excel_data.seek(0)
        response = HttpResponse(excel_data, content_type="application/ms-excel")
        response["Content-Disposition"] = f'attachment; filename=my_timesheet_{date.year}_{date.month}_{user.username}.xlsx'
        return response